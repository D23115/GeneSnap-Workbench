import os
import tempfile
import unittest
from decimal import Decimal
from datetime import date, datetime, timezone
from pathlib import Path

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from PySide6.QtCore import QDate, QItemSelectionModel, Qt
from PySide6.QtWidgets import (
    QApplication,
    QComboBox,
    QDialogButtonBox,
    QScrollArea,
    QTableWidgetItem,
)
from Bio.Seq import Seq
from Bio.SeqRecord import SeqRecord
from Bio import SeqIO
from openpyxl import Workbook

from genesnap_workbench.app.application import (
    GeneSnapApplicationService,
    NewExpressionProjectCommand,
    NewReporterProjectCommand,
    NewShRNAProjectCommand,
    NewSYNProjectCommand,
)
from genesnap_workbench.domain.shrna import (
    BlastScreenStatus,
    ShRNACandidate,
)
from genesnap_workbench.integrations.ncbi_transcripts import TranscriptCandidate
from genesnap_workbench.integrations.shrna_online import ShRNAOnlineDesignResult
from genesnap_workbench.template_engine.workbook_templates import inspect_workbook_template
from genesnap_workbench.app.desktop import (
    ImportExpressionProtocolDialog,
    ImportReporterProtocolDialog,
    ImportWorkbookTemplateDialog,
    MainWindow,
    NewProjectDialog,
    NewExpressionProjectDialog,
    NewReporterProjectDialog,
    NewShRNAProjectDialog,
    NewSYNProjectDialog,
)
from tests.test_syn_design_engine import artificial_sequence
from tests.test_expression_vector_protocol import vector_and_protocol
from tests.test_reporter_exports import export_promoter_sequence
from tests.test_reporter_vector_protocol import vector_and_protocol as reporter_vector_and_protocol


NOW = datetime(2026, 7, 12, 15, 0, tzinfo=timezone.utc)


class DesktopUITests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.qt_app = QApplication.instance() or QApplication([])

    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.service = GeneSnapApplicationService(Path(self.temp_dir.name))
        self.window = MainWindow(
            self.service,
            actor="测试用户",
            today_provider=lambda: date(2026, 7, 12),
        )
        self.addCleanup(self.window.close)

    def create_project(self):
        command = NewSYNProjectCommand(
            project_id="SYN-UI-001",
            target_name="artificial-600",
            raw_sequence=artificial_sequence(600, seed=7),
            input_format="plain",
            linearization_site="EcoRV",
            received_date=date(2026, 7, 12),
            due_date=date(2026, 8, 3),
            actor="测试用户",
            vector_sequence_confirmed=True,
        )
        prepared = self.service.prepare_syn_project(command, created_at=NOW)
        return self.service.save_prepared_syn_project(
            command,
            prepared,
            design_confirmation_reason="已人工复核",
            created_at=NOW,
        )

    def test_main_window_is_project_dashboard_not_landing_page(self):
        self.assertEqual(self.window.windowTitle(), "GeneSnap Workbench")
        self.assertEqual(self.window.project_table.rowCount(), 0)
        self.assertIn("项目号", self.window.project_table.headers())
        self.assertIn("剩余工作日", self.window.project_table.headers())
        self.assertTrue(self.window.new_project_action.isEnabled())
        self.assertTrue(self.window.new_expression_project_action.isEnabled())
        self.assertTrue(self.window.import_expression_protocol_action.isEnabled())
        self.assertTrue(self.window.new_reporter_project_action.isEnabled())
        self.assertTrue(self.window.import_reporter_protocol_action.isEnabled())

    def test_toolbar_shows_action_names_and_new_project_opens_type_selector(self):
        self.assertEqual(
            self.window.main_toolbar.toolButtonStyle(),
            Qt.ToolButtonStyle.ToolButtonTextUnderIcon,
        )
        self.assertEqual(self.window.new_project_action.text(), "新建项目")
        self.assertEqual(self.window.save_location_action.text(), "保存位置")

        dialog = NewProjectDialog(self.service.projects_root, self.window)
        self.addCleanup(dialog.close)
        labels = tuple(dialog.project_type.itemText(i) for i in range(dialog.project_type.count()))
        self.assertEqual(labels, ("shRNA 敲低", "表达类", "GL002 报告载体", "全基因合成"))
        self.assertEqual(Path(dialog.output_root.text()), self.service.projects_root)

    def test_project_intake_dialogs_fit_short_screens_and_keep_forms_scrollable(self):
        dialogs = (
            NewShRNAProjectDialog(
                self.window.calendar,
                self.window,
                service=self.service,
            ),
            NewExpressionProjectDialog(
                self.window.calendar,
                self.service,
                parent=self.window,
            ),
            NewReporterProjectDialog(
                self.window.calendar,
                self.service,
                parent=self.window,
            ),
            NewSYNProjectDialog(self.window.calendar, self.window),
        )
        screen = self.qt_app.primaryScreen().availableGeometry()

        for dialog in dialogs:
            with self.subTest(dialog=type(dialog).__name__):
                self.addCleanup(dialog.close)
                dialog.show()
                self.qt_app.processEvents()

                scroll_area = dialog.findChild(QScrollArea, "projectIntakeScrollArea")
                self.assertIsNotNone(scroll_area)
                self.assertEqual(
                    scroll_area.verticalScrollBarPolicy(),
                    Qt.ScrollBarPolicy.ScrollBarAlwaysOn,
                )
                self.assertLessEqual(dialog.width(), screen.width())
                self.assertLessEqual(dialog.height(), screen.height())

                dialog.resize(560, 480)
                self.qt_app.processEvents()
                self.assertLessEqual(dialog.width(), 560)
                self.assertLessEqual(dialog.height(), 480)
                self.assertEqual(scroll_area.horizontalScrollBar().maximum(), 0)

                buttons = dialog.findChild(QDialogButtonBox)
                self.assertIsNotNone(buttons)
                self.assertTrue(buttons.isVisible())
                button_bottom = buttons.mapTo(dialog, buttons.rect().bottomRight()).y()
                self.assertLessEqual(button_bottom, dialog.contentsRect().bottom())

    def test_project_table_has_an_always_visible_horizontal_scrollbar(self):
        self.window.project_table.setRowCount(1)
        for column in range(self.window.project_table.columnCount()):
            value = f"column-{column}"
            if column >= 9:
                value = "需要横向滚动才能查看的完整项目信息"
            self.window.project_table.setItem(0, column, QTableWidgetItem(value))

        self.window.resize(1040, 680)
        self.window.show()
        self.qt_app.processEvents()

        scrollbar = self.window.project_table.horizontalScrollBar()
        self.assertEqual(
            self.window.project_table.horizontalScrollBarPolicy(),
            Qt.ScrollBarPolicy.ScrollBarAlwaysOn,
        )
        self.assertEqual(
            self.window.project_table.horizontalScrollMode(),
            self.window.project_table.ScrollMode.ScrollPerPixel,
        )
        self.assertTrue(scrollbar.isVisible())
        self.assertGreaterEqual(scrollbar.height(), 18)
        self.assertGreater(scrollbar.maximum(), 0)
        scrollbar.setValue(scrollbar.maximum())
        self.assertEqual(scrollbar.value(), scrollbar.maximum())

    def test_project_dates_show_full_year_and_cannot_precede_received_date(self):
        dialog = NewShRNAProjectDialog(
            self.window.calendar,
            self.window,
            service=self.service,
        )
        self.addCleanup(dialog.close)

        dialog.received_date.setDate(QDate(2026, 7, 13))
        self.qt_app.processEvents()

        self.assertEqual(dialog.received_date.displayFormat(), "yyyy-MM-dd")
        self.assertEqual(dialog.due_date.displayFormat(), "yyyy-MM-dd")
        self.assertEqual(dialog.due_date.minimumDate(), QDate(2026, 7, 13))
        self.assertEqual(dialog.due_date.date(), QDate(2026, 7, 24))
        dialog.due_date.setDate(QDate(2025, 7, 24))
        self.assertGreaterEqual(dialog.due_date.date(), dialog.received_date.date())

    def test_template_mapping_uses_labeled_dropdowns_instead_of_raw_column_numbers(self):
        source = Path(self.temp_dir.name) / "tsingke-primer.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "引物合成订单"
        sheet["A5"] = " *客户姓名："
        sheet["A6"] = " *负责人姓名："
        headers = ("ID", "引物名称", "引物序列(5'to3')", "碱基数", "纯化方法", "OD/管")
        for column, value in enumerate(headers, start=1):
            sheet.cell(12, column).value = value
        workbook.save(source)
        inspected = inspect_workbook_template(source, kind="primer_order")
        dialog = ImportWorkbookTemplateDialog(self.window)
        self.addCleanup(dialog.close)
        dialog._source_path = source
        dialog.source_path.setText(str(source))
        dialog.display_name.setText("擎科引物订购表")

        dialog._populate_inspection(inspected)

        sequence_row = next(
            row
            for row in range(dialog.table_mapping.rowCount())
            if dialog.table_mapping.item(row, 0).data(Qt.ItemDataRole.UserRole) == "sequence"
        )
        sequence_combo = dialog.table_mapping.cellWidget(sequence_row, 1)
        self.assertIsInstance(sequence_combo, QComboBox)
        self.assertEqual(sequence_combo.currentData(), 3)
        self.assertIn("C", sequence_combo.currentText())
        self.assertIn("引物序列", sequence_combo.currentText())

        customer_row = next(
            row
            for row in range(dialog.contact_mapping.rowCount())
            if dialog.contact_mapping.item(row, 0).data(Qt.ItemDataRole.UserRole)
            == "customer_name"
        )
        contact_combo = dialog.contact_mapping.cellWidget(customer_row, 1)
        self.assertIsInstance(contact_combo, QComboBox)
        self.assertEqual(contact_combo.currentData(), "B5")

        _, _, confirmed = dialog.import_data()
        self.assertEqual(confirmed.table_columns["sequence"], 3)
        self.assertEqual(confirmed.contact_cells["customer_name"], "B5")

    def test_detail_tabs_explain_read_only_and_available_interactions(self):
        self.window.tabs.setCurrentIndex(2)
        self.qt_app.processEvents()

        self.assertIn("只读", self.window.detail_hint.text())
        self.assertIn("上方", self.window.detail_hint.text())
        self.assertIn("测序分析", self.window.detail_hint.text())

        self.window.tabs.setCurrentIndex(4)
        self.qt_app.processEvents()
        self.assertIn("双击", self.window.detail_hint.text())

    def test_refresh_select_and_copy_project_row(self):
        self.create_project()
        self.window.refresh_projects()

        self.assertEqual(self.window.project_table.rowCount(), 1)
        project_id_item = self.window.project_table.item(0, 0)
        self.assertEqual(project_id_item.text(), "SYN-UI-001")
        self.window.project_table.selectionModel().select(
            self.window.project_table.model().index(0, 0),
            QItemSelectionModel.SelectionFlag.Select,
        )
        self.window.project_table.copy_selected_cells()
        self.assertEqual(QApplication.clipboard().text(), "SYN-UI-001")

        self.window.project_table.selectRow(0)
        self.window.on_project_selection_changed()
        self.assertIn("SYN-UI-001", self.window.detail_title.text())
        self.assertGreater(self.window.oligo_table.rowCount(), 0)

    def test_context_action_updates_database_and_refreshes_status(self):
        self.create_project()
        self.window.refresh_projects()
        self.window.project_table.setCurrentCell(0, 0)
        self.window.on_project_selection_changed()

        self.assertIn("标记 oligo 已订购", self.window.action_texts())
        button = next(
            item
            for item in self.window.action_buttons
            if item.text() == "标记 oligo 已订购"
        )
        button.click()
        self.qt_app.processEvents()

        stored = self.service.load_project("SYN-UI-001")
        self.assertEqual(stored.snapshot.status, "materials_ordered")
        self.assertIn("标记 oligo 已到货", self.window.action_texts())

    def test_shrna_project_appears_in_same_dashboard_with_targets_and_files(self):
        self.service.create_shrna_project(
            NewShRNAProjectCommand(
                project_id="KD-UI-001",
                gene_symbol="TP53",
                species="human",
                cds_sequence="ATG" * 300,
                candidates=(
                    ShRNACandidate(
                        candidate_id="manual-1",
                        target_sequence="GACTCCAGTGGTAATCTACTG",
                        start_position=None,
                        intrinsic_score=Decimal("0"),
                        source_rank=1,
                        blast_status=BlastScreenStatus.MANUALLY_ACCEPTED,
                    ),
                ),
                target_count=1,
                clones_per_target=5,
                received_date=date(2026, 7, 12),
                due_date=date(2026, 7, 23),
                actor="测试用户",
                vector_sequence_confirmed=True,
            ),
            created_at=NOW,
        )

        self.window.refresh_projects()
        self.window.project_table.setCurrentCell(0, 0)
        self.window.on_project_selection_changed()

        self.assertEqual(self.window.project_table.item(0, 0).text(), "KD-UI-001")
        self.assertEqual(self.window.project_table.item(0, 2).text(), "沉默/敲低类")
        self.assertIn("KD-UI-001", self.window.detail_title.text())
        self.assertEqual(self.window.oligo_table.rowCount(), 2)
        self.assertEqual(self.window.experiment_table.rowCount(), 5)
        self.assertEqual(self.window.artifact_table.rowCount(), 5)
        self.assertIn("标记引物已订购", self.window.action_texts())
        self.assertIn("修正完工日期", self.window.action_texts())

    def test_shrna_dialog_can_generate_ranked_targets_from_cds(self):
        dialog = NewShRNAProjectDialog(self.window.calendar, self.window)
        self.addCleanup(dialog.close)
        dialog.cds_sequence.setPlainText(">NM_TEST\n" + "ACGT" * 180)

        dialog._generate_candidates()

        lines = dialog._target_lines()
        self.assertEqual(len(lines), 3)
        self.assertTrue(dialog.candidate_confirmation.isEnabled())
        dialog.candidate_confirmation.setChecked(True)
        command = dialog.command("test-user")
        self.assertTrue(
            all(
                item.blast_status is BlastScreenStatus.MANUALLY_ACCEPTED
                for item in command.candidates
            ),
        )

    def test_shrna_dialog_keeps_automatic_blast_pass_status(self):
        target = ShRNACandidate(
            candidate_id="broad-1",
            target_sequence="GACTCCAGTGGTAATCTACTG",
            start_position=120,
            intrinsic_score=Decimal("15.0"),
            source_rank=1,
            blast_status=BlastScreenStatus.PASS,
            blast_note="自动 BLAST 通过",
            forward_oligo_sequence=(
                "CCGGGACTCCAGTGGTAATCTACTGCTCGAGCAGTAGATTACCACTGGAGTCTTTTTG"
            ),
            reverse_oligo_sequence=(
                "AATTCAAAAAGACTCCAGTGGTAATCTACTGCTCGAGCAGTAGATTACCACTGGAGTC"
            ),
            oligo_source="broad_gpp",
        )
        dialog = NewShRNAProjectDialog(self.window.calendar, self.window)
        self.addCleanup(dialog.close)
        dialog.target_count.setValue(1)

        dialog._apply_online_design_result(
            ShRNAOnlineDesignResult(
                selected_candidates=(target,),
                candidate_pool=(target,),
                requires_manual_confirmation=False,
                notes=("Broad GPP 返回 1 条候选",),
            ),
        )

        command = dialog.command("test-user")
        self.assertEqual(command.candidates[0].blast_status, BlastScreenStatus.PASS)
        self.assertEqual(dialog.online_result_table.rowCount(), 1)

    def test_shrna_dialog_can_fill_cds_from_transcript_accession(self):
        candidate = TranscriptCandidate(
            accession="NM_TEST.2",
            gene_symbol="TP53",
            description="TP53 mRNA",
            cds_sequence="ATG" + "GCT" * 30 + "TAA",
            protein_id="NP_TEST.1",
            is_mane_select=True,
            is_refseq_select=False,
        )

        class FakeClient:
            def fetch_accession(self, accession):
                self.accession = accession
                return candidate

        client = FakeClient()
        dialog = NewShRNAProjectDialog(
            self.window.calendar,
            self.window,
            ncbi_client=client,
        )
        self.addCleanup(dialog.close)
        dialog.transcript_accession.setText("NM_TEST.2")

        dialog._lookup_transcript()

        self.assertEqual(client.accession, "NM_TEST.2")
        self.assertEqual(dialog.gene_symbol.text(), "TP53")
        self.assertEqual(dialog.cds_sequence.toPlainText(), candidate.cds_sequence)

    def test_shrna_dialog_reuses_saved_primer_and_sequencing_templates(self):
        def save_template(filename, kind, header):
            path = Path(self.temp_dir.name) / filename
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = header
            sheet["B1"] = "引物序列" if kind == "primer_order" else "测序引物"
            workbook.save(path)
            inspected = inspect_workbook_template(path, kind=kind)
            return self.service.save_workbook_template(
                path,
                display_name=filename.removesuffix(".xlsx"),
                inspected=inspected,
            )

        primer = save_template("擎科.xlsx", "primer_order", "引物名称")
        sequencing = save_template("华大.xlsx", "sequencing_order", "样本名称")
        dialog = NewShRNAProjectDialog(
            self.window.calendar,
            self.window,
            service=self.service,
        )
        self.addCleanup(dialog.close)
        dialog.primer_template_combo.setCurrentIndex(1)
        dialog.sequencing_template_combo.setCurrentIndex(1)
        dialog.targets.setPlainText("GACTCCAGTGGTAATCTACTG")

        command = dialog.command("test-user")

        self.assertEqual(command.primer_template_id, primer.template_id)
        self.assertEqual(command.sequencing_template_id, sequencing.template_id)
        self.assertEqual(command.primer_vendor_name, "擎科")
        self.assertEqual(command.sequencing_vendor_name, "华大")

    def test_expression_project_appears_with_construct_primers_clones_and_files(self):
        vector, protocol = vector_and_protocol()
        self.service.create_expression_project(
            NewExpressionProjectCommand(
                project_id="OE-UI-001",
                gene_symbol="TP53",
                species="human",
                source_cds="ATG" + "GCT" * 120 + "TAA",
                construct_lines=("FL", "1-80aa"),
                received_date=date(2026, 7, 12),
                due_date=date(2026, 7, 23),
                actor="测试用户",
                vector=vector,
                protocol=protocol,
                clones_per_construct=5,
            ),
            created_at=NOW,
        )

        self.window.refresh_projects()
        self.window.project_table.setCurrentCell(0, 0)
        self.window.on_project_selection_changed()

        self.assertEqual(self.window.project_table.item(0, 0).text(), "OE-UI-001")
        self.assertEqual(self.window.project_table.item(0, 2).text(), "表达类")
        self.assertIn("OE-UI-001", self.window.detail_title.text())
        self.assertEqual(self.window.oligo_table.rowCount(), 4)
        self.assertEqual(self.window.experiment_table.rowCount(), 10)
        self.assertEqual(self.window.artifact_table.rowCount(), 6)
        self.assertIn("标记引物已订购", self.window.action_texts())

    def test_expression_dialog_uses_saved_profile_and_one_line_constructs(self):
        vector, protocol = vector_and_protocol()
        saved = self.service.save_expression_profile(vector, protocol)
        dialog = NewExpressionProjectDialog(self.window.calendar, self.service)
        self.addCleanup(dialog.close)
        dialog.project_id.setText("OE-DIALOG-001")
        dialog.gene_symbol.setText("TP53")
        dialog.cds_sequence.setPlainText("ATG" + "GCT" * 120 + "TAA")
        dialog.construct_lines.setPlainText("FL\n1-80aa")

        command = dialog.command("测试用户")

        self.assertEqual(dialog.profile_combo.currentData(), saved.profile_id)
        self.assertEqual(command.vector, vector)
        self.assertEqual(command.protocol, protocol)
        self.assertEqual(command.construct_lines, ("FL", "1-80aa"))

    def test_expression_dialog_reuses_ncbi_transcript_lookup(self):
        candidate = TranscriptCandidate(
            accession="NM_TEST.2",
            gene_symbol="TP53",
            description="TP53 mRNA",
            cds_sequence="ATG" + "GCT" * 30 + "TAA",
            protein_id="NP_TEST.1",
            is_mane_select=True,
            is_refseq_select=False,
        )

        class FakeClient:
            def fetch_accession(self, accession):
                return candidate

        dialog = NewExpressionProjectDialog(
            self.window.calendar,
            self.service,
            ncbi_client=FakeClient(),
        )
        self.addCleanup(dialog.close)
        dialog.transcript_accession.setText("NM_TEST.2")

        dialog._lookup_transcript()

        self.assertEqual(dialog.cds_sequence.toPlainText(), candidate.cds_sequence)

    def test_expression_warning_results_expose_manual_review_actions(self):
        vector, protocol = vector_and_protocol()
        self.service.create_expression_project(
            NewExpressionProjectCommand(
                project_id="OE-REVIEW-001",
                gene_symbol="TP53",
                species="human",
                source_cds="ATG" + "GCT" * 120 + "TAA",
                construct_lines=("FL",),
                received_date=date(2026, 7, 12),
                due_date=date(2026, 7, 23),
                actor="测试用户",
                vector=vector,
                protocol=protocol,
            ),
            created_at=NOW,
        )
        self.service.analyze_expression_sequencing(
            "OE-REVIEW-001",
            actor="测试用户",
            analyzed_at=NOW,
        )

        self.window.refresh_projects()
        self.window.project_table.setCurrentCell(0, 0)
        self.window.on_project_selection_changed()

        self.assertIn("确认可用", self.window.action_texts())
        self.assertIn("确认不可用", self.window.action_texts())

    def test_import_protocol_dialog_reads_genbank_and_builds_valid_profile(self):
        vector, protocol = vector_and_protocol()
        vector_path = Path(self.temp_dir.name) / "artificial-vector.gb"
        record = SeqRecord(Seq(vector.sequence), id="artificial", name="artificial")
        record.annotations["molecule_type"] = "DNA"
        record.annotations["topology"] = "circular"
        SeqIO.write(record, vector_path, "genbank")
        dialog = ImportExpressionProtocolDialog()
        self.addCleanup(dialog.close)
        dialog.load_vector_path(vector_path)
        dialog.display_name.setText(protocol.display_name)
        dialog.protocol_version_id.setText(protocol.protocol_version_id)
        dialog.left_boundary.setValue(protocol.left_boundary)
        dialog.right_boundary.setValue(protocol.right_boundary)
        dialog.left_homology.setText(protocol.left_primer_homology)
        dialog.right_homology.setText(protocol.right_primer_homology)

        imported_vector, imported_protocol = dialog.profile()

        self.assertEqual(imported_vector.sequence, vector.sequence)
        self.assertEqual(imported_protocol.vector_checksum, vector.normalized_circular_sha256)
        self.assertEqual(imported_protocol.left_boundary, protocol.left_boundary)

    def test_reporter_project_appears_with_constructs_primers_clones_and_files(self):
        vector, protocol = reporter_vector_and_protocol()
        self.service.create_reporter_project(
            NewReporterProjectCommand(
                project_id="RPT-UI-001",
                gene_symbol="SGK1",
                species="human",
                promoter_sequence=export_promoter_sequence(),
                construct_lines=("WT", "P1500", "P1000", "P500"),
                mutation_definitions=(),
                received_date=date(2026, 7, 12),
                due_date=date(2026, 7, 23),
                actor="测试用户",
                vector=vector,
                protocol=protocol,
            ),
            created_at=NOW,
        )

        self.window.refresh_projects()
        self.window.project_table.setCurrentCell(0, 0)
        self.window.on_project_selection_changed()

        self.assertEqual(self.window.project_table.item(0, 0).text(), "RPT-UI-001")
        self.assertEqual(self.window.project_table.item(0, 2).text(), "报告/检测类")
        self.assertIn("RPT-UI-001", self.window.detail_title.text())
        self.assertEqual(self.window.oligo_table.rowCount(), 8)
        self.assertEqual(self.window.experiment_table.rowCount(), 20)
        self.assertEqual(self.window.artifact_table.rowCount(), 8)
        self.assertIn("标记引物已订购", self.window.action_texts())

    def test_reporter_dialog_uses_saved_profile(self):
        vector, protocol = reporter_vector_and_protocol()
        saved = self.service.save_reporter_profile(vector, protocol)
        dialog = NewReporterProjectDialog(self.window.calendar, self.service)
        self.addCleanup(dialog.close)
        dialog.project_id.setText("RPT-DIALOG-001")
        dialog.gene_symbol.setText("SGK1")
        dialog.promoter_sequence.setPlainText(export_promoter_sequence())
        dialog.construct_lines.setPlainText("WT\nP1500\nP1000\nP500")

        command = dialog.command("测试用户")

        self.assertEqual(dialog.profile_combo.currentData(), saved.profile_id)
        self.assertEqual(command.vector, vector)
        self.assertEqual(command.protocol, protocol)
        self.assertEqual(command.construct_lines, ("WT", "P1500", "P1000", "P500"))


if __name__ == "__main__":
    unittest.main()
