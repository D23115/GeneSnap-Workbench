import tempfile
import unittest
from dataclasses import replace
from datetime import date, datetime, timezone
from pathlib import Path

from Bio import SeqIO
from docx import Document
from openpyxl import Workbook, load_workbook

from genesnap_workbench.domain.expression import ExpressionDesignInput
from genesnap_workbench.project_workflow.project_folders import create_project_folder
from genesnap_workbench.sequence_core.expression import create_expression_design
from genesnap_workbench.template_engine.expression_exports import (
    ExpressionExportError,
    export_expression_bundle,
)
from genesnap_workbench.template_engine.workbook_templates import (
    LocalWorkbookTemplateStore,
    inspect_workbook_template,
)
from genesnap_workbench.vector_library.expression import (
    apply_expression_protocol,
    expression_rules_from_protocol,
)
from tests.test_expression_vector_protocol import vector_and_protocol


NOW = datetime(2026, 7, 12, 23, 30, tzinfo=timezone.utc)


def make_design_and_vector_result(construct_lines=("FL", "1-80aa", "Δ81-100")):
    vector, protocol = vector_and_protocol()
    design = create_expression_design(
        ExpressionDesignInput(
            project_id="OE-EXPORT-001",
            gene_symbol="TP53",
            species="human",
            source_cds="ATG" + "GCT" * 120 + "TAA",
            construct_lines=construct_lines,
        ),
        expression_rules_from_protocol(protocol),
        design_version_id="OE-EXPORT-001-v1",
        created_at=NOW,
    )
    return design, apply_expression_protocol(design, vector, protocol)


class ExpressionExportTests(unittest.TestCase):
    @staticmethod
    def _primer_template(root: Path):
        source = root / "primer-template.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "引物合成订单"
        sheet["A1"] = "下单日期"
        sheet.merge_cells("B1:C1")
        for column, value in enumerate(
            ("引物名称", "引物序列", "基因ID", "转录本号", "PCR产物长度"),
            start=1,
        ):
            sheet.cell(4, column).value = value
        workbook.save(source)
        store = LocalWorkbookTemplateStore(root / "templates")
        profile = store.save_import(
            source,
            display_name="供应商引物表",
            inspected=inspect_workbook_template(source, kind="primer_order"),
        )
        return store, profile

    def test_three_constructs_export_six_primers_and_fifteen_clones(self):
        with tempfile.TemporaryDirectory() as directory:
            workspace = create_project_folder(
                Path(directory),
                project_id="OE-EXPORT-001",
                target_name="TP53",
                folder_suffix="OE",
            )
            design, vector_result = make_design_and_vector_result()

            bundle = export_expression_bundle(
                design,
                vector_result,
                workspace,
                generated_at=NOW,
                primer_order_date=date(2026, 7, 13),
                sequencing_order_date=date(2026, 7, 15),
                clones_per_construct=5,
                primer_vendor_name="标准",
                sequencing_vendor_name="标准",
            )

            primer_book = load_workbook(bundle.path_for("primer_order_xlsx"))
            sequencing_book = load_workbook(bundle.path_for("sequencing_order_xlsx"))
            primer_rows = [
                row
                for row in primer_book.active.iter_rows(min_row=7, values_only=True)
                if row[0]
            ]
            sequencing_rows = [
                row
                for row in sequencing_book.active.iter_rows(min_row=7, values_only=True)
                if row[0]
            ]

            self.assertEqual(len(primer_rows), 6)
            self.assertEqual(len(sequencing_rows), 15)
            self.assertEqual(sequencing_rows[0][0], "TP53-FL-1")
            self.assertEqual(sequencing_rows[-1][0], "TP53-Δ81-100-5")
            self.assertTrue(all(row[1] == "Nanopore" for row in sequencing_rows))

    def test_bundle_contains_editable_report_and_map_per_construct(self):
        with tempfile.TemporaryDirectory() as directory:
            workspace = create_project_folder(
                Path(directory),
                project_id="OE-EXPORT-001",
                target_name="TP53",
                folder_suffix="OE",
            )
            design, vector_result = make_design_and_vector_result(("FL", "1-80aa"))

            bundle = export_expression_bundle(
                design,
                vector_result,
                workspace,
                generated_at=NOW,
                primer_order_date=date(2026, 7, 13),
                sequencing_order_date=date(2026, 7, 15),
            )

            report = Document(bundle.path_for("design_report_docx"))
            maps = bundle.paths_for("expected_plasmid_genbank")
            map_records = tuple(SeqIO.read(path, "genbank") for path in maps)

            self.assertIn("TP53", "\n".join(item.text for item in report.paragraphs))
            self.assertEqual(len(maps), 2)
            self.assertEqual(
                tuple(len(item.seq) for item in map_records),
                tuple(len(item.expected_plasmid_sequence) for item in vector_result.construct_plans),
            )
            self.assertTrue(all(item.path.exists() for item in bundle.artifacts))

    def test_unconfirmed_mutation_design_cannot_export_formal_orders(self):
        design, vector_result = make_design_and_vector_result(("A2G",))
        with tempfile.TemporaryDirectory() as directory:
            workspace = create_project_folder(
                Path(directory),
                project_id="OE-EXPORT-001",
                target_name="TP53",
                folder_suffix="OE",
            )

            with self.assertRaises(ExpressionExportError):
                export_expression_bundle(
                    design,
                    vector_result,
                    workspace,
                    generated_at=NOW,
                    primer_order_date=date(2026, 7, 13),
                    sequencing_order_date=date(2026, 7, 15),
                )

    def test_long_construct_exports_four_fragment_primers(self):
        vector, protocol = vector_and_protocol()
        long_design = create_expression_design(
            ExpressionDesignInput(
                project_id="OE-LONG-EXPORT",
                gene_symbol="LONG1",
                species="human",
                source_cds="ATG" + "GCT" * 2400 + "TAA",
                construct_lines=("FL",),
            ),
            expression_rules_from_protocol(protocol),
            design_version_id="OE-LONG-EXPORT-v1",
            created_at=NOW,
        )
        long_result = apply_expression_protocol(long_design, vector, protocol)
        with tempfile.TemporaryDirectory() as directory:
            workspace = create_project_folder(
                Path(directory),
                project_id="OE-LONG-EXPORT",
                target_name="LONG1",
                folder_suffix="OE",
            )

            bundle = export_expression_bundle(
                long_design,
                long_result,
                workspace,
                generated_at=NOW,
                primer_order_date=date(2026, 7, 13),
                sequencing_order_date=date(2026, 7, 15),
            )

            book = load_workbook(bundle.path_for("primer_order_xlsx"))
            rows = [
                row
                for row in book.active.iter_rows(min_row=7, values_only=True)
                if row[0]
            ]
            self.assertEqual(len(rows), 4)
            self.assertEqual(tuple(row[0] for row in rows), (
                "LONG1-FL-P1-F",
                "LONG1-FL-P1-R",
                "LONG1-FL-P2-F",
                "LONG1-FL-P2-R",
            ))

    def test_vendor_template_uses_each_primers_actual_pcr_fragment_length(self):
        vector, protocol = vector_and_protocol()
        design = create_expression_design(
            ExpressionDesignInput(
                project_id="OE-TEMPLATE-001",
                gene_symbol="TP53",
                species="human",
                source_cds="ATG" + "GCT" * 2401 + "TAA",
                construct_lines=("FL",),
                transcript_accession="NM_000546.6",
            ),
            expression_rules_from_protocol(protocol),
            design_version_id="OE-TEMPLATE-001-v1",
            created_at=NOW,
        )
        design = replace(design, gene_id="7157")
        vector_result = apply_expression_protocol(design, vector, protocol)
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            workspace = create_project_folder(
                root,
                project_id="OE-TEMPLATE-001",
                target_name="TP53",
                folder_suffix="OE",
            )
            store, profile = self._primer_template(root)

            bundle = export_expression_bundle(
                design,
                vector_result,
                workspace,
                generated_at=NOW,
                primer_order_date=date(2026, 7, 15),
                sequencing_order_date=date(2026, 7, 16),
                workbook_template_store=store,
                primer_template_id=profile.template_id,
            )

            sheet = load_workbook(bundle.path_for("primer_order_xlsx"))["引物合成订单"]
            rows = [row for row in sheet.iter_rows(min_row=5, values_only=True) if row[0]]
            fragment_lengths = {
                fragment.fragment_no: len(fragment.sequence)
                for fragment in design.constructs[0].fragments
            }
            primer_fragment_numbers = {
                primer.name: primer.fragment_no
                for primer in vector_result.construct_plans[0].primers
            }
            self.assertEqual(sheet["B1"].value.date(), NOW.date())
            self.assertTrue(all(row[2] == "7157" for row in rows))
            self.assertTrue(all(row[3] == "NM_000546.6" for row in rows))
            self.assertEqual(
                {row[0]: row[4] for row in rows},
                {
                    name: fragment_lengths[fragment_no]
                    for name, fragment_no in primer_fragment_numbers.items()
                },
            )


if __name__ == "__main__":
    unittest.main()
