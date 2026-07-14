import tempfile
import unittest
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path

from Bio import SeqIO
from docx import Document
from openpyxl import Workbook, load_workbook

from genesnap_workbench.domain.shrna import (
    BlastScreenStatus,
    ShRNACandidate,
    ShRNADesignInput,
)
from genesnap_workbench.project_workflow.project_folders import create_project_folder
from genesnap_workbench.sequence_core.shrna import create_shrna_design
from genesnap_workbench.template_engine.shrna_exports import export_shrna_bundle
from genesnap_workbench.template_engine.workbook_templates import (
    LocalWorkbookTemplateStore,
    inspect_workbook_template,
)
from genesnap_workbench.vector_library.starters import load_public_plko1_puro_starter


NOW = datetime(2026, 7, 12, 18, 30, tzinfo=timezone.utc)


def make_design(target_count: int):
    vector, protocol = load_public_plko1_puro_starter(user_confirmed=True)
    design_input = ShRNADesignInput(
        project_id=f"KD-EXPORT-{target_count}",
        gene_symbol="TP53",
        species="human",
        cds_sequence="ATG" * 300,
        vector_protocol_version_id=protocol.protocol_version_id,
        transcript_accession="NM_000546.6",
        gene_id="7157",
        target_count=target_count,
    )
    sequences = (
        "GACTCCAGTGGTAATCTACTG",
        "CCTGAGGTTGGCTCTGACTGT",
        "TGGATGATTTGATGCTGTCCC",
    )
    candidates = tuple(
        ShRNACandidate(
            candidate_id=f"candidate-{number}",
            target_sequence=sequences[number - 1],
            start_position=number * 200,
            intrinsic_score=Decimal(str(10 - number)),
            source_rank=number,
            blast_status=BlastScreenStatus.PASS,
        )
        for number in range(1, target_count + 1)
    )
    return create_shrna_design(
        design_input,
        candidates,
        vector,
        protocol,
        design_version_id=f"KD-EXPORT-{target_count}-v1",
        created_at=NOW,
    )


class ShRNAExportTests(unittest.TestCase):
    @staticmethod
    def _primer_template(root: Path):
        source = root / "primer-template.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "引物合成订单"
        sheet["A1"] = "订购日期"
        sheet.merge_cells("B1:C1")
        for column, value in enumerate(
            ("引物名称", "引物序列", "Gene ID", "NM号", "扩增产物长度"),
            start=1,
        ):
            sheet.cell(4, column).value = value
        workbook.save(source)
        store = LocalWorkbookTemplateStore(root / "templates")
        profile = store.save_import(
            source,
            display_name="供应商引物表",
            inspected=inspect_workbook_template(source, kind="primer_order"),
        )
        return store, profile

    def test_three_targets_export_six_oligos_and_fifteen_clones(self):
        with tempfile.TemporaryDirectory() as directory:
            workspace = create_project_folder(
                Path(directory),
                project_id="KD-EXPORT-3",
                target_name="TP53",
                folder_suffix="KD",
            )
            bundle = export_shrna_bundle(
                make_design(3),
                workspace,
                generated_at=NOW,
                primer_order_date=date(2026, 7, 12),
                sequencing_order_date=date(2026, 7, 14),
                primer_vendor_name="标准",
                sequencing_vendor_name="标准",
            )

            primer_book = load_workbook(bundle.path_for("primer_order_xlsx"))
            sequencing_book = load_workbook(bundle.path_for("sequencing_order_xlsx"))

            primer_rows = list(primer_book.active.iter_rows(min_row=7, values_only=True))
            sequencing_rows = list(
                sequencing_book.active.iter_rows(min_row=7, values_only=True),
            )
            self.assertEqual(len([row for row in primer_rows if row[0]]), 6)
            self.assertEqual(len([row for row in sequencing_rows if row[0]]), 15)
            self.assertEqual(sequencing_rows[0][0], "TP53-1-1")
            self.assertEqual(sequencing_rows[-1][0], "TP53-3-5")
            self.assertTrue(all(row[1] == "U6" for row in sequencing_rows if row[0]))

    def test_bundle_contains_editable_report_and_one_expected_map_per_target(self):
        with tempfile.TemporaryDirectory() as directory:
            workspace = create_project_folder(
                Path(directory),
                project_id="KD-EXPORT-1",
                target_name="TP53",
                folder_suffix="KD",
            )
            design = make_design(1)
            bundle = export_shrna_bundle(
                design,
                workspace,
                generated_at=NOW,
                primer_order_date=date(2026, 7, 12),
                sequencing_order_date=date(2026, 7, 14),
            )

            report = Document(bundle.path_for("design_report_docx"))
            maps = bundle.paths_for("expected_plasmid_genbank")
            record = SeqIO.read(maps[0], "genbank")

            self.assertIn("TP53", "\n".join(item.text for item in report.paragraphs))
            self.assertEqual(len(maps), 1)
            self.assertEqual(len(record.seq), len(design.plasmid_simulations[0].expected_plasmid_sequence))
            self.assertTrue(all(item.path.exists() for item in bundle.artifacts))

    def test_vendor_template_receives_gene_transcript_oligo_length_and_order_date(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            workspace = create_project_folder(
                root,
                project_id="KD-EXPORT-1",
                target_name="TP53",
                folder_suffix="KD",
            )
            store, profile = self._primer_template(root)
            design = make_design(1)

            bundle = export_shrna_bundle(
                design,
                workspace,
                generated_at=NOW,
                primer_order_date=date(2026, 7, 15),
                sequencing_order_date=date(2026, 7, 16),
                workbook_template_store=store,
                primer_template_id=profile.template_id,
            )

            sheet = load_workbook(bundle.path_for("primer_order_xlsx"))["引物合成订单"]
            rows = list(sheet.iter_rows(min_row=5, max_row=6, values_only=True))
            self.assertEqual(sheet["B1"].value.date(), NOW.date())
            self.assertTrue(all(row[2] == "7157" for row in rows))
            self.assertTrue(all(row[3] == "NM_000546.6" for row in rows))
            self.assertEqual(
                tuple(row[4] for row in rows),
                (
                    len(design.targets[0].oligos.forward_sequence),
                    len(design.targets[0].oligos.reverse_sequence),
                ),
            )


if __name__ == "__main__":
    unittest.main()
