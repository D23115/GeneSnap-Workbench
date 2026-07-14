import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from genesnap_workbench.vector_library.comparison_report import (
    analyze_vector_libraries,
    write_vector_comparison_report,
)


class VectorComparisonReportTests(unittest.TestCase):
    def test_finds_renamed_exact_vector_and_best_near_match(self):
        with TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            internal_root = root / "internal"
            public_root = root / "public"
            internal_root.mkdir()
            public_root.mkdir()

            exact = "AACCGTTAAGTCGGATCCGATGCTAGCTAGGCTA"
            rotated = exact[11:] + exact[:11]
            public_near = "TTGACCGATCGTACGATGGCCTAACCGGTTAACCGG"
            internal_near = public_near[:15] + "T" + public_near[16:]
            (internal_root / "内部A.fasta").write_text(
                f">internal_a\n{exact}\n", encoding="ascii"
            )
            (public_root / "公共别名.fasta").write_text(
                f">public_alias\n{rotated}\n", encoding="ascii"
            )
            (internal_root / "内部B.fasta").write_text(
                f">internal_b\n{internal_near}\n", encoding="ascii"
            )
            (public_root / "公共B.fasta").write_text(
                f">public_b\n{public_near}\n", encoding="ascii"
            )

            report = analyze_vector_libraries(
                internal_root,
                public_root,
                internal_extensions={".fasta"},
                public_extensions={".fasta"},
                k=5,
                sketch_size=100,
                candidate_count=2,
            )

            self.assertEqual(len(report.exact_pairs), 1)
            self.assertEqual(report.exact_pairs[0].internal.file_name, "内部A.fasta")
            self.assertEqual(report.exact_pairs[0].public.file_name, "公共别名.fasta")
            matches = {item.internal.file_name: item for item in report.best_matches}
            self.assertEqual(matches["内部A.fasta"].classification, "完全一致")
            self.assertEqual(matches["内部B.fasta"].public.file_name, "公共B.fasta")
            self.assertEqual(matches["内部B.fasta"].comparison.edit_distance, 1)

    def test_writes_readable_multisheet_workbook(self):
        with TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            internal_root = root / "internal"
            public_root = root / "public"
            internal_root.mkdir()
            public_root.mkdir()
            sequence = "AACCGTTAAGTCGGATCCGATGCTAGCTAGGCTA"
            (internal_root / "internal.fasta").write_text(
                f">internal\n{sequence}\n", encoding="ascii"
            )
            (public_root / "public.fasta").write_text(
                f">public\n{sequence}\n", encoding="ascii"
            )
            report = analyze_vector_libraries(
                internal_root,
                public_root,
                internal_extensions={".fasta"},
                public_extensions={".fasta"},
                k=5,
                sketch_size=100,
            )
            output = root / "comparison.xlsx"

            write_vector_comparison_report(report, output)

            workbook = load_workbook(output, read_only=True)
            self.assertIn("总览", workbook.sheetnames)
            self.assertIn("完全一致_跨库", workbook.sheetnames)
            self.assertIn("内部载体最佳匹配", workbook.sheetnames)
            self.assertIn("差异区段", workbook.sheetnames)
            workbook.close()

    def test_does_not_expand_unreliable_match_into_difference_rows(self):
        with TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            internal_root = root / "internal"
            public_root = root / "public"
            internal_root.mkdir()
            public_root.mkdir()
            (internal_root / "unrelated.fasta").write_text(
                ">internal\n" + "A" * 60 + "C" * 60 + "\n", encoding="ascii"
            )
            (public_root / "public.fasta").write_text(
                ">public\n" + "G" * 60 + "A" * 60 + "\n", encoding="ascii"
            )
            report = analyze_vector_libraries(
                internal_root,
                public_root,
                internal_extensions={".fasta"},
                public_extensions={".fasta"},
                k=5,
                sketch_size=100,
            )
            output = root / "comparison.xlsx"

            write_vector_comparison_report(report, output)

            workbook = load_workbook(output, read_only=True)
            difference_row_count = workbook["差异区段"].max_row
            workbook.close()
            self.assertEqual(difference_row_count, 1)


if __name__ == "__main__":
    unittest.main()
