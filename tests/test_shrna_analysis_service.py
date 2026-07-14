import tempfile
import unittest
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from genesnap_workbench.app.application import (
    GeneSnapApplicationService,
    NewShRNAProjectCommand,
)
from genesnap_workbench.domain.shrna import BlastScreenStatus, ShRNACandidate


NOW = datetime(2026, 7, 12, 20, 0, tzinfo=timezone.utc)


class ShRNAAnalysisServiceTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.service = GeneSnapApplicationService(Path(self.temp_dir.name))
        self.stored = self.service.create_shrna_project(
            NewShRNAProjectCommand(
                project_id="KD-ANALYSIS-001",
                gene_symbol="TP53",
                species="human",
                cds_sequence="ATG" * 300,
                candidates=(
                    ShRNACandidate(
                        candidate_id="manual-1",
                        target_sequence="GACTCCAGTGGTAATCTACTG",
                        start_position=None,
                        intrinsic_score=Decimal("0"),
                        source_rank=1,
                        blast_status=BlastScreenStatus.MANUALLY_ACCEPTED,
                    ),
                ),
                target_count=1,
                clones_per_target=5,
                received_date=date(2026, 7, 12),
                due_date=date(2026, 7, 23),
                actor="tester",
                vector_sequence_confirmed=True,
            ),
            created_at=NOW,
        )

    def test_scan_matches_names_inside_vendor_filenames_and_records_missing_clones(self):
        sequencing_dir = self.stored.project_folder / "03_sequencing" / "vendor-return"
        sequencing_dir.mkdir()
        expected = self.stored.design.targets[0].oligos.forward_sequence
        (sequencing_dir / "prefix_TP53-1-1_U6_vendor.seq").write_text(
            "A" * 40 + expected + "C" * 50,
            encoding="ascii",
        )
        (sequencing_dir / "second_TP53-1-1_U6_vendor.txt").write_text(
            "A" * 40 + expected + "C" * 50,
            encoding="ascii",
        )
        mutated = list(expected)
        mutated[18] = "A" if mutated[18] != "A" else "C"
        (sequencing_dir / "TP53-1-2_U6_vendor.seq").write_text(
            "A" * 40 + "".join(mutated) + "C" * 50,
            encoding="ascii",
        )
        (sequencing_dir / "unmatched_vendor_file.seq").write_text(
            "ACGT" * 80,
            encoding="ascii",
        )

        outcome = self.service.analyze_shrna_sequencing(
            self.stored.project_id,
            actor="tester",
            analyzed_at=NOW,
        )

        latest = {item.clone_name: item for item in outcome.project.snapshot.clone_results}
        self.assertEqual(latest["TP53-1-1"].status, "pass")
        self.assertIn("2 个测序文件判读一致", latest["TP53-1-1"].reason)
        self.assertEqual(latest["TP53-1-2"].status, "fail")
        self.assertEqual(latest["TP53-1-3"].status, "warning")
        self.assertEqual(len(outcome.unmatched_files), 1)
        self.assertEqual(outcome.project.snapshot.status, "analysis_completed")
        report = load_workbook(outcome.analysis_report)
        self.assertEqual(report.active.max_row, 11)
        artifacts = self.service.shrna_repository.list_artifacts(self.stored.project_id)
        self.assertEqual(artifacts[-1].artifact_type, "sequencing_analysis_xlsx")

        reviewed = self.service.confirm_shrna_clone_review(
            self.stored.project_id,
            clone_name="TP53-1-3",
            usable=False,
            note="未找到有效测序文件，按不可用处理",
            actor="tester",
            reviewed_at=NOW,
        )
        latest = {item.clone_name: item for item in reviewed.snapshot.clone_results}
        self.assertEqual(latest["TP53-1-3"].manual_review_status, "unusable")
        self.assertFalse(latest["TP53-1-3"].manually_confirmed_usable)


if __name__ == "__main__":
    unittest.main()
