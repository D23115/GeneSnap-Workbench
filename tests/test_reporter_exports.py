import tempfile
import unittest
from dataclasses import replace
from datetime import date, datetime, timezone
from pathlib import Path

from Bio import SeqIO
from docx import Document
from openpyxl import Workbook, load_workbook

from genesnap_workbench.domain.reporter import ReporterDesignInput
from genesnap_workbench.project_workflow.project_folders import create_project_folder
from genesnap_workbench.sequence_core.reporter import create_reporter_design
from genesnap_workbench.template_engine.reporter_exports import (
    ReporterExportError,
    export_reporter_bundle,
)
from genesnap_workbench.template_engine.workbook_templates import (
    LocalWorkbookTemplateStore,
    inspect_workbook_template,
)
from genesnap_workbench.vector_library.reporter import apply_reporter_protocol
from tests.test_reporter_engine import promoter_sequence
from tests.test_reporter_vector_protocol import vector_and_protocol


NOW = datetime(2026, 7, 13, 3, 0, tzinfo=timezone.utc)


def export_promoter_sequence():
    sequence = list(promoter_sequence(2000))
    for start, marker in (
        (0, "A" * 20),
        (500, "C" * 20),
        (1000, "G" * 20),
        (1500, "T" * 20),
    ):
        sequence[start : start + 20] = marker
    return "".join(sequence)


def make_design_and_result(*, mutation=False):
    vector, protocol = vector_and_protocol()
    design = create_reporter_design(
        ReporterDesignInput(
            project_id="RPT-EXPORT-001",
            gene_symbol="SGK1",
            species="human",
            promoter_sequence=export_promoter_sequence(),
            mutation_definitions=("mut1:101-104=TTTT",) if mutation else (),
            construct_lines=("mut1",) if mutation else ("WT", "P1500", "P1000", "P500"),
        ),
        protocol_version_id=protocol.protocol_version_id,
        design_version_id="RPT-EXPORT-001-v1",
        created_at=NOW,
    )
    return design, apply_reporter_protocol(design, vector, protocol)


class ReporterExportTests(unittest.TestCase):
    @staticmethod
    def _primer_template(root: Path):
        source = root / "primer-template.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "引物合成订单"
        sheet["A1"] = "订单日期"
        sheet.merge_cells("B1:C1")
        for column, value in enumerate(
            ("引物名称", "引物序列", "Gene ID", "NM号", "扩增产物长度"),
            start=1,
        ):
            sheet.cell(4, column).value = value
        workbook.save(source)
        store = LocalWorkbookTemplateStore(root / "templates")
        profile = store.save_import(
            source,
            display_name="供应商引物表",
            inspected=inspect_workbook_template(source, kind="primer_order"),
        )
        return store, profile

    def test_four_deletions_export_five_unique_primers_and_twenty_clones(self):
        with tempfile.TemporaryDirectory() as directory:
            workspace = create_project_folder(
                Path(directory),
                project_id="RPT-EXPORT-001",
                target_name="SGK1",
                folder_suffix="RPT",
            )
            design, result = make_design_and_result()

            bundle = export_reporter_bundle(
                design,
                result,
                workspace,
                generated_at=NOW,
                primer_order_date=date(2026, 7, 13),
                sequencing_order_date=date(2026, 7, 15),
                clones_per_construct=5,
            )

            primer_book = load_workbook(bundle.path_for("primer_order_xlsx"))
            sequencing_book = load_workbook(bundle.path_for("sequencing_order_xlsx"))
            primer_rows = [
                row for row in primer_book.active.iter_rows(min_row=7, values_only=True) if row[0]
            ]
            sequencing_rows = [
                row
                for row in sequencing_book.active.iter_rows(min_row=7, values_only=True)
                if row[0]
            ]
            self.assertEqual(len(primer_rows), 5)
            self.assertEqual(tuple(row[0] for row in primer_rows).count("SGK1-P-R"), 1)
            self.assertEqual(len(sequencing_rows), 20)
            self.assertEqual(sequencing_rows[0][0], "SGK1-promoter-WT-1")
            self.assertTrue(all(row[1] == "Nanopore" for row in sequencing_rows))

    def test_bundle_contains_editable_report_and_map_per_construct(self):
        with tempfile.TemporaryDirectory() as directory:
            workspace = create_project_folder(
                Path(directory),
                project_id="RPT-EXPORT-001",
                target_name="SGK1",
                folder_suffix="RPT",
            )
            design, result = make_design_and_result()

            bundle = export_reporter_bundle(
                design,
                result,
                workspace,
                generated_at=NOW,
                primer_order_date=date(2026, 7, 13),
                sequencing_order_date=date(2026, 7, 15),
            )

            report = Document(bundle.path_for("design_report_docx"))
            maps = bundle.paths_for("expected_plasmid_genbank")
            records = tuple(SeqIO.read(path, "genbank") for path in maps)
            self.assertIn("SGK1", "\n".join(item.text for item in report.paragraphs))
            self.assertEqual(len(records), 4)
            self.assertEqual(
                tuple(len(item.seq) for item in records),
                tuple(len(item.expected_plasmid_sequence) for item in result.construct_plans),
            )

    def test_unconfirmed_mutation_cannot_export(self):
        design, result = make_design_and_result(mutation=True)
        with tempfile.TemporaryDirectory() as directory:
            workspace = create_project_folder(
                Path(directory),
                project_id="RPT-EXPORT-001",
                target_name="SGK1",
                folder_suffix="RPT",
            )
            with self.assertRaises(ReporterExportError):
                export_reporter_bundle(
                    design,
                    result,
                    workspace,
                    generated_at=NOW,
                    primer_order_date=date(2026, 7, 13),
                    sequencing_order_date=date(2026, 7, 15),
                )

    def test_shared_primer_uses_sorted_readable_product_lengths_in_vendor_template(self):
        design, result = make_design_and_result()
        design = replace(
            design,
            gene_id="6446",
            transcript_accession="NM_001143676.3",
        )
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            workspace = create_project_folder(
                root,
                project_id="RPT-EXPORT-001",
                target_name="SGK1",
                folder_suffix="RPT",
            )
            store, profile = self._primer_template(root)

            bundle = export_reporter_bundle(
                design,
                result,
                workspace,
                generated_at=NOW,
                primer_order_date=date(2026, 7, 15),
                sequencing_order_date=date(2026, 7, 16),
                workbook_template_store=store,
                primer_template_id=profile.template_id,
            )

            sheet = load_workbook(bundle.path_for("primer_order_xlsx"))["引物合成订单"]
            rows = [row for row in sheet.iter_rows(min_row=5, values_only=True) if row[0]]
            by_name = {row[0]: row for row in rows}
            self.assertEqual(sheet["B1"].value.date(), NOW.date())
            self.assertTrue(all(row[2] == "6446" for row in rows))
            self.assertTrue(all(row[3] == "NM_001143676.3" for row in rows))
            self.assertEqual(
                {name: row[4] for name, row in by_name.items()},
                {
                    "SGK1-P2000-F": 2000,
                    "SGK1-P1500-F": 1500,
                    "SGK1-P1000-F": 1000,
                    "SGK1-P500-F": 500,
                    # 共享 R 引物的不同产物长度去重后升序、用逗号分隔。
                    "SGK1-P-R": "500, 1000, 1500, 2000",
                },
            )


if __name__ == "__main__":
    unittest.main()
