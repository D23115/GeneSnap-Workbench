import json
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from genesnap_workbench.template_engine.workbook_templates import (
    ContactProfile,
    LocalContactProfileStore,
    LocalWorkbookTemplateStore,
    inspect_workbook_template,
    workbook_mapping_choices,
)


class WorkbookTemplateTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.root = Path(self.temp_dir.name)
        self.source = self.root / "vendor-primer.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "订购单"
        sheet["A1"] = "客户姓名"
        sheet["C1"] = "客户单位"
        sheet["A4"] = "引物名称"
        sheet["B4"] = "引物序列（5'-3'）"
        sheet["C4"] = "纯化方式"
        sheet["D4"] = "备注"
        sheet["A5"] = "示例行"
        workbook.save(self.source)

    def test_inspection_guesses_table_and_contact_mapping(self):
        inspected = inspect_workbook_template(self.source, kind="primer_order")

        self.assertEqual(inspected.sheet_name, "订购单")
        self.assertEqual(inspected.header_row, 4)
        self.assertEqual(inspected.data_start_row, 5)
        self.assertEqual(inspected.table_columns["primer_name"], 1)
        self.assertEqual(inspected.table_columns["sequence"], 2)
        self.assertEqual(inspected.table_columns["purification"], 3)
        self.assertEqual(inspected.contact_cells["customer_name"], "B1")
        self.assertEqual(inspected.contact_cells["organization"], "D1")

    def test_inspection_recognizes_tsingke_style_primer_headers_and_contacts(self):
        source = self.root / "tsingke-primer.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "引物合成订单"
        sheet["A3"] = "订购日期"
        sheet.merge_cells("B3:C3")
        sheet["A5"] = " *客户姓名："
        sheet["A6"] = " *负责人姓名："
        sheet["A7"] = " *客户单位："
        sheet["A9"] = " *客户地址："
        sheet["A10"] = " *联系电话："
        sheet["A11"] = " *客户Email："
        headers = (
            "ID",
            "引物名称",
            "引物序列(5'to3')",
            "碱基数",
            "纯化方法",
            "OD/管",
            "管数",
            "5'修饰",
            "3'修饰",
            "Gene ID",
            "NM号",
            "扩增产物长度",
            "中间修饰",
            "备注",
        )
        for column, value in enumerate(headers, start=1):
            sheet.cell(12, column).value = value
        workbook.save(source)

        inspected = inspect_workbook_template(source, kind="primer_order")

        self.assertEqual(inspected.sheet_name, "引物合成订单")
        self.assertEqual(inspected.header_row, 12)
        self.assertEqual(
            inspected.table_columns,
            {
                "primer_name": 2,
                "sequence": 3,
                "length": 4,
                "purification": 5,
                "scale": 6,
                "gene_id": 10,
                "transcript_accession": 11,
                "product_length": 12,
                "note": 14,
            },
        )
        self.assertEqual(inspected.document_cells, {"order_date": "B3"})
        self.assertEqual(
            inspected.contact_cells,
            {
                "customer_name": "B5",
                "responsible_name": "B6",
                "organization": "B7",
                "address": "B9",
                "phone": "B10",
                "email": "B11",
            },
        )

    def test_inspection_recognizes_alternate_row_and_document_aliases(self):
        source = self.root / "alternate-aliases.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "订单"
        sheet["A1"] = "下单日期"
        sheet.merge_cells("B1:C1")
        headers = ("引物名称", "基因ID", "转录本号", "PCR产物长度")
        for column, value in enumerate(headers, start=1):
            sheet.cell(3, column).value = value
        workbook.save(source)

        inspected = inspect_workbook_template(source, kind="primer_order")

        self.assertEqual(
            inspected.table_columns,
            {
                "primer_name": 1,
                "gene_id": 2,
                "transcript_accession": 3,
                "product_length": 4,
            },
        )
        self.assertEqual(inspected.document_cells, {"order_date": "B1"})

    def test_inspection_recognizes_tsingke_expanded_product_length_header(self):
        source = self.root / "tsingke-expanded-product-length.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "引物合成订单"
        sheet["A2"] = "引物名称"
        sheet["B2"] = "扩展产物长度"
        workbook.save(source)

        inspected = inspect_workbook_template(source, kind="primer_order")

        self.assertEqual(
            inspected.table_columns,
            {"primer_name": 1, "product_length": 2},
        )

    def test_sequencing_inspection_recognizes_shared_metadata_fields(self):
        source = self.root / "sequencing-metadata.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "送测单"
        headers = ("样本名称", "Gene ID", "转录本号", "PCR产物长度")
        for column, value in enumerate(headers, start=1):
            sheet.cell(2, column).value = value
        workbook.save(source)

        inspected = inspect_workbook_template(source, kind="sequencing_order")

        self.assertEqual(
            inspected.table_columns,
            {
                "sample_name": 1,
                "gene_id": 2,
                "transcript_accession": 3,
                "product_length": 4,
            },
        )

    def test_merged_targets_use_anchor_for_mapping_rendering_and_historical_clear(self):
        source = self.root / "merged-vendor.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "引物合成订单"
        sheet["A1"] = "订单日期"
        sheet.merge_cells("B1:C1")
        sheet["A2"] = "客户姓名"
        sheet.merge_cells("B2:C2")
        sheet["A4"] = "引物名称"
        sheet["B4"] = "引物序列"
        sheet["C4"] = "备注"
        sheet.merge_cells("C4:D4")
        sheet.merge_cells("C5:D5")
        sheet["C5"] = "旧备注"
        sheet.merge_cells("C6:D6")
        sheet["C6"] = "应清空的历史行"
        workbook.save(source)

        inspected = inspect_workbook_template(source, kind="primer_order")
        inspected.document_cells["order_date"] = "C1"
        inspected.contact_cells["customer_name"] = "C2"
        inspected.table_columns["note"] = 4
        choices = workbook_mapping_choices(
            source,
            sheet_name=inspected.sheet_name,
            header_row=inspected.header_row,
        )
        candidate_coordinates = {coordinate for coordinate, _ in choices.contact_cells}
        candidate_columns = {column for column, _ in choices.table_columns}
        self.assertNotIn("C1", candidate_coordinates)
        self.assertNotIn("C2", candidate_coordinates)
        self.assertNotIn(4, candidate_columns)

        store = LocalWorkbookTemplateStore(self.root / "merged-templates")
        saved = store.save_import(source, display_name="合并单元格模板", inspected=inspected)
        self.assertEqual(saved.document_cells["order_date"], "B1")
        self.assertEqual(saved.contact_cells["customer_name"], "B2")

        output = self.root / "merged-filled.xlsx"
        store.render(
            saved.template_id,
            records=(
                {
                    "primer_name": "TP53-F",
                    "sequence": "ACGT",
                    "note": "新备注",
                },
            ),
            contact=ContactProfile(customer_name="测试客户"),
            document_values={"order_date": date(2026, 7, 14)},
            output_path=output,
        )

        rendered = load_workbook(output, data_only=False)["引物合成订单"]
        self.assertIsInstance(rendered["B1"].value, datetime)
        self.assertEqual(rendered["B1"].value.date(), date(2026, 7, 14))
        self.assertEqual(rendered["B2"].value, "测试客户")
        self.assertEqual(rendered["C5"].value, "新备注")
        self.assertIsNone(rendered["C6"].value)

    def test_loading_legacy_non_anchor_mappings_normalizes_profile(self):
        source = self.root / "legacy-merged.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "订单"
        sheet["A1"] = "订单日期"
        sheet.merge_cells("B1:C1")
        sheet["A2"] = "客户姓名"
        sheet.merge_cells("B2:C2")
        sheet["A4"] = "引物名称"
        sheet["B4"] = "引物序列"
        sheet["C4"] = "备注"
        sheet.merge_cells("C5:D5")
        workbook.save(source)

        store = LocalWorkbookTemplateStore(self.root / "legacy-merged-templates")
        inspected = inspect_workbook_template(source, kind="primer_order")
        saved = store.save_import(source, display_name="旧合并模板", inspected=inspected)
        metadata = (
            self.root
            / "legacy-merged-templates"
            / saved.template_id
            / "profile.json"
        )
        payload = json.loads(metadata.read_text(encoding="utf-8"))
        payload["document_cells"] = {"order_date": "C1"}
        payload["contact_cells"] = {"customer_name": "C2"}
        payload["table_columns"]["note"] = 4
        metadata.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )

        loaded = store.load_profile(saved.template_id)

        self.assertEqual(loaded.document_cells["order_date"], "B1")
        self.assertEqual(loaded.contact_cells["customer_name"], "B2")
        self.assertEqual(loaded.table_columns["note"], 3)

    def test_profile_without_document_cells_remains_readable(self):
        store = LocalWorkbookTemplateStore(self.root / "legacy-templates")
        inspected = inspect_workbook_template(self.source, kind="primer_order")
        saved = store.save_import(self.source, display_name="旧模板", inspected=inspected)
        metadata = self.root / "legacy-templates" / saved.template_id / "profile.json"
        payload = json.loads(metadata.read_text(encoding="utf-8"))
        payload.pop("document_cells", None)
        metadata.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )

        loaded = store.load_profile(saved.template_id)

        self.assertEqual(loaded.document_cells, {})

    def test_saved_mapping_reopens_and_renders_without_reconfirmation(self):
        store = LocalWorkbookTemplateStore(self.root / "templates")
        inspected = inspect_workbook_template(self.source, kind="primer_order")
        saved = store.save_import(
            self.source,
            display_name="擎科引物订购表",
            inspected=inspected,
        )
        reopened = LocalWorkbookTemplateStore(self.root / "templates")

        self.assertEqual(reopened.list_profiles("primer_order"), (saved,))
        loaded = reopened.load_profile(saved.template_id)
        self.assertEqual(loaded.table_columns, inspected.table_columns)

        output = self.root / "filled.xlsx"
        reopened.render(
            saved.template_id,
            records=(
                {
                    "primer_name": "TP53-1-F",
                    "sequence": "ACGTACGT",
                    "purification": "PAGE",
                },
                {
                    "primer_name": "TP53-1-R",
                    "sequence": "TGCATGCA",
                },
            ),
            contact=ContactProfile(
                customer_name="示例用户",
                organization="示例单位",
            ),
            output_path=output,
        )

        workbook = load_workbook(output, data_only=False)
        sheet = workbook["订购单"]
        self.assertEqual(sheet["B1"].value, "示例用户")
        self.assertEqual(sheet["D1"].value, "示例单位")
        self.assertEqual(sheet["A5"].value, "TP53-1-F")
        self.assertEqual(sheet["B6"].value, "TGCATGCA")
        self.assertIsNone(sheet["C6"].value)

    def test_unmapped_optional_fields_do_not_block_rendering(self):
        store = LocalWorkbookTemplateStore(self.root / "templates")
        inspected = inspect_workbook_template(self.source, kind="primer_order")
        inspected.table_columns.pop("purification")
        saved = store.save_import(
            self.source,
            display_name="精简模板",
            inspected=inspected,
        )

        store.render(
            saved.template_id,
            records=({"primer_name": "P1", "sequence": "ACGT", "scale": "5 OD"},),
            contact=ContactProfile(customer_name="测试"),
            output_path=self.root / "optional.xlsx",
        )

        self.assertTrue((self.root / "optional.xlsx").exists())

    def test_single_default_contact_profile_can_be_changed(self):
        store = LocalContactProfileStore(self.root / "contact.json")
        self.assertEqual(store.load(), ContactProfile())

        store.save(ContactProfile(customer_name="示例用户", phone="00000000000"))

        reopened = LocalContactProfileStore(self.root / "contact.json")
        self.assertEqual(reopened.load().customer_name, "示例用户")
        self.assertEqual(reopened.load().phone, "00000000000")


if __name__ == "__main__":
    unittest.main()
