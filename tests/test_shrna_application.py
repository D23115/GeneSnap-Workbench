import tempfile
import unittest
from dataclasses import replace
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook

from genesnap_workbench.app.application import (
    GeneSnapApplicationService,
    NewShRNAProjectCommand,
)
from genesnap_workbench.domain.shrna import (
    BlastScreenStatus,
    ShRNACandidate,
)
from genesnap_workbench.template_engine.workbook_templates import (
    ContactProfile,
    inspect_workbook_template,
)


NOW = datetime(2026, 7, 12, 19, 0, tzinfo=timezone.utc)


class ShRNAApplicationTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.service = GeneSnapApplicationService(Path(self.temp_dir.name))

    def command(self):
        return NewShRNAProjectCommand(
            project_id="KD-APP-001",
            gene_symbol="TP53",
            species="human",
            cds_sequence="ATG" * 300,
            candidates=(
                ShRNACandidate(
                    candidate_id="candidate-1",
                    target_sequence="GACTCCAGTGGTAATCTACTG",
                    start_position=120,
                    intrinsic_score=Decimal("9.1"),
                    source_rank=1,
                    blast_status=BlastScreenStatus.PASS,
                ),
            ),
            target_count=1,
            clones_per_target=5,
            received_date=date(2026, 7, 12),
            due_date=date(2026, 7, 23),
            actor="tester",
            vector_sequence_confirmed=True,
        )

    def test_create_shrna_project_saves_database_and_complete_output_bundle(self):
        stored = self.service.create_shrna_project(self.command(), created_at=NOW)

        self.assertEqual(stored.snapshot.status, "design_completed")
        self.assertTrue(stored.project_folder.name.endswith("-KD"))
        self.assertTrue((stored.project_folder / "03_sequencing").is_dir())
        artifacts = self.service.shrna_repository.list_artifacts(stored.project_id)
        self.assertEqual(len(artifacts), 5)
        self.assertTrue(all(item.path.exists() for item in artifacts))
        self.assertEqual(len(stored.design.targets[0].clone_names), 5)

        reopened = GeneSnapApplicationService(Path(self.temp_dir.name))
        restored = reopened.load_shrna_project(stored.project_id)
        self.assertEqual(restored.design, stored.design)
        self.assertEqual(restored.snapshot, stored.snapshot)

    def test_rejects_due_date_before_received_date_without_creating_project(self):
        command = replace(self.command(), due_date=date(2025, 7, 24))

        with self.assertRaisesRegex(ValueError, "不能早于接收日期"):
            self.service.create_shrna_project(command, created_at=NOW)

        self.assertEqual(self.service.shrna_repository.list_projects(), ())

    def test_due_date_correction_keeps_original_date_and_appends_history(self):
        stored = self.service.create_shrna_project(self.command(), created_at=NOW)

        corrected = self.service.adjust_molecular_due_date(
            stored.project_id,
            workflow_type="shrna_knockdown",
            new_due_date=date(2026, 7, 24),
            note="修正录入年份",
            actor="tester",
            occurred_at=NOW,
        )

        self.assertEqual(corrected.due_date, date(2026, 7, 23))
        self.assertEqual(corrected.snapshot.effective_due_date, date(2026, 7, 24))
        self.assertEqual(corrected.snapshot.status_history[-1].event_type, "adjust_due_date")
        self.assertIn("修正录入年份", corrected.snapshot.status_history[-1].note)

    def test_unconfirmed_public_vector_is_rejected_before_folder_creation(self):
        command = self.command()
        command = replace(command, vector_sequence_confirmed=False)

        with self.assertRaisesRegex(ValueError, "protocol"):
            self.service.create_shrna_project(command, created_at=NOW)
        self.assertFalse(self.service.projects_root.exists())

    def test_project_can_render_primer_rows_into_saved_vendor_template(self):
        template_path = Path(self.temp_dir.name) / "vendor.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "引物订单"
        sheet["A1"] = "客户姓名"
        sheet["A3"] = "引物名称"
        sheet["B3"] = "引物序列（5'-3'）"
        workbook.save(template_path)
        inspected = inspect_workbook_template(template_path, kind="primer_order")
        profile = self.service.save_workbook_template(
            template_path,
            display_name="擎科引物订购表",
            inspected=inspected,
        )
        self.service.save_contact_profile(ContactProfile(customer_name="示例用户"))

        command = replace(
            self.command(),
            primer_vendor_name="擎科",
            primer_template_id=profile.template_id,
        )
        stored = self.service.create_shrna_project(command, created_at=NOW)

        primer_artifact = next(
            item
            for item in self.service.shrna_repository.list_artifacts(stored.project_id)
            if item.artifact_type == "primer_order_xlsx"
        )
        filled = load_workbook(primer_artifact.path)
        sheet = filled["引物订单"]
        self.assertEqual(sheet["B1"].value, "示例用户")
        self.assertEqual(sheet["A4"].value, "TP53-1-F")
        self.assertEqual(sheet["B5"].value, stored.design.targets[0].oligos.reverse_sequence)


if __name__ == "__main__":
    unittest.main()
