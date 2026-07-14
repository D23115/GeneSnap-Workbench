"""Sequencing forms generated for later add-on and post-rework rounds."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .workbook_templates import ContactProfile, LocalWorkbookTemplateStore


def export_sequencing_form(
    records: tuple[dict[str, object], ...],
    output_path: Path,
    *,
    template_store: LocalWorkbookTemplateStore,
    contact_profile: ContactProfile,
    template_id: str | None = None,
) -> Path:
    output = Path(output_path)
    if template_id:
        return template_store.render(
            template_id,
            records=records,
            contact=contact_profile,
            output_path=output,
        )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "送测"
    headers = ("样本名称", "测序引物", "目标基因", "克隆号", "测序方式", "备注")
    for column_no, header in enumerate(headers, start=1):
        cell = sheet.cell(1, column_no, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F6F62")
    for row_no, record in enumerate(records, start=2):
        values = (
            record.get("sample_name"),
            record.get("primer_name"),
            record.get("gene_symbol"),
            record.get("clone_no"),
            record.get("method"),
            record.get("note"),
        )
        for column_no, value in enumerate(values, start=1):
            sheet.cell(row_no, column_no, value)
    sheet.column_dimensions["A"].width = 34
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output
