"""Reusable vendor workbook inspection, persistence, and rendering."""

from __future__ import annotations

from copy import copy
from dataclasses import asdict, dataclass, field
import hashlib
import json
from pathlib import Path
import re
import shutil
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


_SUPPORTED_KINDS = frozenset({"primer_order", "sequencing_order"})


def _normalize_label(value: object) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[\s\-_（）()'\"：:，,。.／/\\*＊]+", "", text)


def _aliases(*values: str) -> frozenset[str]:
    return frozenset(_normalize_label(value) for value in values)


_TABLE_ALIASES = {
    "primer_order": {
        "primer_name": _aliases("引物名称", "引物名", "oligo名称", "序列名称", "name"),
        "sequence": _aliases(
            "引物序列",
            "引物序列（5'-3'）",
            "引物序列(5'-3')",
            "引物序列(5'to3')",
            "引物序列5to3",
            "序列5'-3'",
            "碱基序列",
            "sequence",
        ),
        "gene_symbol": _aliases("基因", "基因名", "目标基因", "gene"),
        "gene_id": _aliases("Gene ID", "基因ID"),
        "transcript_accession": _aliases("NM号", "转录本号"),
        "product_length": _aliases(
            "扩增产物长度", "PCR产物长度", "扩展产物长度"
        ),
        "direction": _aliases("方向", "引物方向", "direction"),
        "length": _aliases("长度", "碱基数", "长度nt", "length"),
        "purification": _aliases(
            "纯化",
            "纯化方式",
            "纯化方法",
            "纯化级别",
            "purification",
        ),
        "scale": _aliases(
            "合成规模",
            "合成量",
            "需求量",
            "OD/管",
            "OD数",
            "合成OD",
            "scale",
        ),
        "note": _aliases("备注", "特殊说明", "note"),
    },
    "sequencing_order": {
        "sample_name": _aliases("样本名称", "样品名称", "送测名称", "管号", "sample"),
        "primer_name": _aliases("测序引物", "引物名称", "通用引物", "primer"),
        "gene_symbol": _aliases("基因", "基因名", "目标基因", "gene"),
        "gene_id": _aliases("Gene ID", "基因ID"),
        "transcript_accession": _aliases("NM号", "转录本号"),
        "product_length": _aliases(
            "扩增产物长度", "PCR产物长度", "扩展产物长度"
        ),
        "clone_no": _aliases("克隆号", "克隆编号", "clone"),
        "method": _aliases("测序方式", "测序类型", "测序方法", "method"),
        "note": _aliases("备注", "特殊说明", "note"),
    },
}

_CONTACT_ALIASES = {
    "customer_name": _aliases("客户姓名", "客户名称", "联系人", "送样人"),
    "responsible_name": _aliases("负责人姓名", "负责人", "销售负责人"),
    "organization": _aliases("客户单位", "单位名称", "所在单位", "公司名称"),
    "phone": _aliases("联系电话", "手机号码", "手机号", "电话"),
    "email": _aliases("电子邮箱", "邮箱", "客户Email", "email", "e-mail"),
    "address": _aliases("收货地址", "联系地址", "客户地址", "地址"),
    "customer_id": _aliases("客户编号", "客户代码", "账号", "客户ID"),
}

_DOCUMENT_ALIASES = {
    "order_date": _aliases("订购日期", "下单日期", "订单日期"),
}


def table_field_names(kind: str) -> tuple[str, ...]:
    if kind not in _SUPPORTED_KINDS:
        raise ValueError("不支持的模板类型")
    return tuple(_TABLE_ALIASES[kind])


def contact_field_names() -> tuple[str, ...]:
    return tuple(_CONTACT_ALIASES)


@dataclass(frozen=True, slots=True)
class ContactProfile:
    customer_name: str = ""
    responsible_name: str = ""
    organization: str = ""
    phone: str = ""
    email: str = ""
    address: str = ""
    customer_id: str = ""


@dataclass(slots=True)
class WorkbookTemplateInspection:
    kind: str
    sheet_name: str
    header_row: int
    data_start_row: int
    table_columns: dict[str, int]
    contact_cells: dict[str, str]
    document_cells: dict[str, str] = field(default_factory=dict)


@dataclass(frozen=True, slots=True)
class WorkbookMappingChoices:
    table_columns: tuple[tuple[int, str], ...]
    contact_cells: tuple[tuple[str, str], ...]


@dataclass(frozen=True, slots=True)
class WorkbookTemplateProfile:
    template_id: str
    display_name: str
    kind: str
    workbook_filename: str
    workbook_sha256: str
    sheet_name: str
    header_row: int
    data_start_row: int
    table_columns: dict[str, int]
    contact_cells: dict[str, str]
    document_cells: dict[str, str] = field(default_factory=dict)


def _match_field(kind: str, value: object) -> str | None:
    normalized = _normalize_label(value)
    for field_name, aliases in _TABLE_ALIASES[kind].items():
        if normalized in aliases:
            return field_name
    return None


def _merged_anchor_cell(sheet: Any, row: int, column: int) -> Any:
    for merged_range in sheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= column <= merged_range.max_col
        ):
            return sheet.cell(merged_range.min_row, merged_range.min_col)
    return sheet.cell(row, column)


def _merged_anchor_for_coordinate(sheet: Any, coordinate: str) -> Any:
    cell = sheet[coordinate]
    return _merged_anchor_cell(sheet, cell.row, cell.column)


def _normalize_cell_mapping(
    sheet: Any,
    mapping: dict[str, str],
) -> dict[str, str]:
    return {
        field_name: _merged_anchor_for_coordinate(sheet, coordinate).coordinate
        for field_name, coordinate in mapping.items()
    }


def _normalize_table_mapping(
    sheet: Any,
    mapping: dict[str, int],
    *,
    data_start_row: int,
) -> dict[str, int]:
    return {
        field_name: _merged_anchor_cell(
            sheet, data_start_row, column_no
        ).column
        for field_name, column_no in mapping.items()
    }


def inspect_workbook_template(path: Path, *, kind: str) -> WorkbookTemplateInspection:
    source = Path(path)
    if kind not in _SUPPORTED_KINDS:
        raise ValueError("模板类型必须是 primer_order 或 sequencing_order")
    if source.suffix.lower() != ".xlsx":
        raise ValueError("MVP 目前只支持 .xlsx 模板")
    workbook = load_workbook(source, data_only=False)

    best: tuple[int, str, int, dict[str, int]] | None = None
    for sheet in workbook.worksheets:
        for row_no in range(1, min(sheet.max_row, 40) + 1):
            mapping: dict[str, int] = {}
            for column_no in range(1, min(sheet.max_column, 60) + 1):
                field_name = _match_field(kind, sheet.cell(row_no, column_no).value)
                if field_name and field_name not in mapping:
                    mapping[field_name] = column_no
            candidate = (len(mapping), sheet.title, row_no, mapping)
            if best is None or candidate[0] > best[0]:
                best = candidate
    if best is None or best[0] == 0:
        raise ValueError("没有识别到可映射的订购/送测表头")

    _, sheet_name, header_row, table_columns = best
    sheet = workbook[sheet_name]
    contact_cells: dict[str, str] = {}
    document_cells: dict[str, str] = {}
    for row_no in range(1, min(sheet.max_row, 40) + 1):
        for column_no in range(1, min(sheet.max_column, 40) + 1):
            normalized = _normalize_label(sheet.cell(row_no, column_no).value)
            for field_name, aliases in _CONTACT_ALIASES.items():
                if field_name not in contact_cells and normalized in aliases:
                    contact_cells[field_name] = _merged_anchor_cell(
                        sheet, row_no, column_no + 1
                    ).coordinate
            for field_name, aliases in _DOCUMENT_ALIASES.items():
                if field_name not in document_cells and normalized in aliases:
                    document_cells[field_name] = _merged_anchor_cell(
                        sheet, row_no, column_no + 1
                    ).coordinate

    return WorkbookTemplateInspection(
        kind=kind,
        sheet_name=sheet_name,
        header_row=header_row,
        data_start_row=header_row + 1,
        table_columns=table_columns,
        contact_cells=contact_cells,
        document_cells=document_cells,
    )


def workbook_mapping_choices(
    path: Path,
    *,
    sheet_name: str,
    header_row: int,
) -> WorkbookMappingChoices:
    """Return labeled dropdown choices for a previously inspected workbook."""
    workbook = load_workbook(Path(path), data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表不存在：{sheet_name}")
        sheet = workbook[sheet_name]
        if not 1 <= header_row <= sheet.max_row:
            raise ValueError("表头行超出工作表范围")

        table_choices = []
        for column_no in range(1, min(sheet.max_column, 60) + 1):
            cell = sheet.cell(header_row, column_no)
            anchor = _merged_anchor_cell(sheet, header_row, column_no)
            if anchor.coordinate != cell.coordinate:
                continue
            raw_label = cell.value
            label = str(raw_label or "（无表头）").strip().replace("\n", " ")
            table_choices.append(
                (column_no, f"{get_column_letter(column_no)} · {label}"),
            )

        contact_choices = []
        seen_cells: set[str] = set()
        last_contact_row = min(max(header_row - 1, 1), 40)
        for row_no in range(1, last_contact_row + 1):
            for column_no in range(1, min(sheet.max_column, 40) + 1):
                cell = sheet.cell(row_no, column_no)
                anchor = _merged_anchor_cell(sheet, row_no, column_no)
                if anchor.coordinate != cell.coordinate:
                    continue
                raw_label = cell.value
                if raw_label in (None, ""):
                    continue
                target = _merged_anchor_cell(
                    sheet, row_no, column_no + 1
                ).coordinate
                if target in seen_cells:
                    continue
                seen_cells.add(target)
                label = str(raw_label).strip().replace("\n", " ")
                contact_choices.append((target, f"{target} · {label} 右侧"))
        return WorkbookMappingChoices(
            table_columns=tuple(table_choices),
            contact_cells=tuple(contact_choices),
        )
    finally:
        workbook.close()


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _safe_id(value: str) -> str:
    normalized = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff_-]+", "-", value.strip()).strip("-")
    return normalized or "template"


class LocalWorkbookTemplateStore:
    def __init__(self, root: Path) -> None:
        self.root = Path(root)
        self.root.mkdir(parents=True, exist_ok=True)

    def save_import(
        self,
        source_path: Path,
        *,
        display_name: str,
        inspected: WorkbookTemplateInspection,
    ) -> WorkbookTemplateProfile:
        source = Path(source_path)
        if inspected.kind not in _SUPPORTED_KINDS:
            raise ValueError("不支持的模板类型")
        if not display_name.strip():
            raise ValueError("模板名称不能为空")
        source_sha256 = _sha256_file(source)
        template_id = f"{_safe_id(display_name)}-{source_sha256[:10]}"
        profile_dir = self.root / template_id
        profile_dir.mkdir(parents=True, exist_ok=True)
        workbook_filename = "template.xlsx"
        copied_path = profile_dir / workbook_filename
        shutil.copy2(source, copied_path)
        workbook = load_workbook(source, data_only=False)
        try:
            sheet = workbook[inspected.sheet_name]
            table_columns = _normalize_table_mapping(
                sheet,
                inspected.table_columns,
                data_start_row=inspected.data_start_row,
            )
            contact_cells = _normalize_cell_mapping(sheet, inspected.contact_cells)
            document_cells = _normalize_cell_mapping(sheet, inspected.document_cells)
        finally:
            workbook.close()
        profile = WorkbookTemplateProfile(
            template_id=template_id,
            display_name=display_name.strip(),
            kind=inspected.kind,
            workbook_filename=workbook_filename,
            workbook_sha256=source_sha256,
            sheet_name=inspected.sheet_name,
            header_row=inspected.header_row,
            data_start_row=inspected.data_start_row,
            table_columns=table_columns,
            contact_cells=contact_cells,
            document_cells=document_cells,
        )
        (profile_dir / "profile.json").write_text(
            json.dumps(asdict(profile), ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        return profile

    def load_profile(self, template_id: str) -> WorkbookTemplateProfile:
        profile_dir = self.root / template_id
        data = json.loads((profile_dir / "profile.json").read_text(encoding="utf-8"))
        data["table_columns"] = {
            key: int(value) for key, value in data["table_columns"].items()
        }
        data.setdefault("document_cells", {})
        workbook_path = profile_dir / data["workbook_filename"]
        if _sha256_file(workbook_path) != data["workbook_sha256"]:
            raise ValueError("模板文件校验失败，可能已被外部修改")
        workbook = load_workbook(workbook_path, data_only=False)
        try:
            sheet = workbook[data["sheet_name"]]
            data["table_columns"] = _normalize_table_mapping(
                sheet,
                data["table_columns"],
                data_start_row=int(data["data_start_row"]),
            )
            data["contact_cells"] = _normalize_cell_mapping(
                sheet, data["contact_cells"]
            )
            data["document_cells"] = _normalize_cell_mapping(
                sheet, data["document_cells"]
            )
        finally:
            workbook.close()
        return WorkbookTemplateProfile(**data)

    def list_profiles(self, kind: str | None = None) -> tuple[WorkbookTemplateProfile, ...]:
        profiles: list[WorkbookTemplateProfile] = []
        for metadata in sorted(self.root.glob("*/profile.json")):
            profile = self.load_profile(metadata.parent.name)
            if kind is None or profile.kind == kind:
                profiles.append(profile)
        return tuple(sorted(profiles, key=lambda item: (item.display_name, item.template_id)))

    def render(
        self,
        template_id: str,
        *,
        records: tuple[dict[str, Any], ...],
        contact: ContactProfile,
        output_path: Path,
        document_values: dict[str, Any] | None = None,
    ) -> Path:
        profile = self.load_profile(template_id)
        workbook_path = self.root / template_id / profile.workbook_filename
        workbook = load_workbook(workbook_path, data_only=False)
        sheet = workbook[profile.sheet_name]

        contact_values = asdict(contact)
        for field_name, cell_coordinate in profile.contact_cells.items():
            value = contact_values.get(field_name, "")
            if value:
                _merged_anchor_for_coordinate(sheet, cell_coordinate).value = value

        values = document_values or {}
        for field_name, cell_coordinate in profile.document_cells.items():
            if field_name not in values:
                continue
            value = values[field_name]
            if value is not None and value != "":
                _merged_anchor_for_coordinate(sheet, cell_coordinate).value = value

        final_row = profile.data_start_row + max(len(records), 1) - 1
        clear_through = max(sheet.max_row, final_row)
        cleared_cells: set[str] = set()
        for row_no in range(profile.data_start_row, clear_through + 1):
            for column_no in profile.table_columns.values():
                cell = _merged_anchor_cell(sheet, row_no, column_no)
                if cell.coordinate in cleared_cells:
                    continue
                cell.value = None
                cleared_cells.add(cell.coordinate)

        style_row = profile.data_start_row
        for record_index, record in enumerate(records):
            row_no = profile.data_start_row + record_index
            for field_name, column_no in profile.table_columns.items():
                source_cell = _merged_anchor_cell(sheet, style_row, column_no)
                target_cell = _merged_anchor_cell(sheet, row_no, column_no)
                if target_cell.coordinate != source_cell.coordinate:
                    target_cell._style = copy(source_cell._style)
                    target_cell.number_format = source_cell.number_format
                    target_cell.alignment = copy(source_cell.alignment)
                    target_cell.protection = copy(source_cell.protection)
                target_cell.value = record.get(field_name)

        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)
        return output


class LocalContactProfileStore:
    """Persist the single editable workstation contact profile."""

    def __init__(self, path: Path) -> None:
        self.path = Path(path)

    def load(self) -> ContactProfile:
        if not self.path.exists():
            return ContactProfile()
        data = json.loads(self.path.read_text(encoding="utf-8"))
        known = {field_name: data.get(field_name, "") for field_name in ContactProfile.__slots__}
        return ContactProfile(**known)

    def save(self, profile: ContactProfile) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        temporary = self.path.with_suffix(self.path.suffix + ".tmp")
        temporary.write_text(
            json.dumps(asdict(profile), ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        temporary.replace(self.path)
