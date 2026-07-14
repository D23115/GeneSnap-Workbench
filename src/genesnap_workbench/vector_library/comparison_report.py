"""Library scanning and Excel reporting for plasmid sequence comparisons."""

from __future__ import annotations

import csv
import hashlib
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .comparison import (
    CircularComparison,
    canonical_circular_sequence,
    canonical_kmer_sketch,
    classify_similarity,
    compare_circular_sequences,
    infer_internal_sequence_role,
    read_sequence_file,
)


DEFAULT_INTERNAL_EXTENSIONS = {".dna", ".gb", ".gbk", ".fasta", ".fa", ".fna"}
DEFAULT_PUBLIC_EXTENSIONS = {".dna", ".gb", ".gbk", ".fasta", ".fa", ".fna"}


@dataclass(frozen=True)
class LibrarySequence:
    library: str
    file_name: str
    display_name: str
    path: Path
    relative_path: str
    category: str
    role: str
    sequence: str
    format_name: str
    topology: str
    feature_labels: tuple[str, ...]
    ambiguous_base_count: int
    canonical_hash: str
    sketch: tuple[int, ...]

    @property
    def length(self) -> int:
        return len(self.sequence)


@dataclass(frozen=True)
class ParseFailure:
    library: str
    path: Path
    error: str


@dataclass(frozen=True)
class ExactPair:
    internal: LibrarySequence
    public: LibrarySequence


@dataclass(frozen=True)
class BestMatch:
    internal: LibrarySequence
    public: LibrarySequence
    comparison: CircularComparison
    classification: str
    sketch_similarity_percent: float
    exact_public_aliases: tuple[LibrarySequence, ...]


@dataclass(frozen=True)
class DuplicateGroup:
    library: str
    canonical_hash: str
    records: tuple[LibrarySequence, ...]


@dataclass(frozen=True)
class VectorComparisonReport:
    internal_root: Path
    public_root: Path
    generated_at: datetime
    internal_records: tuple[LibrarySequence, ...]
    public_records: tuple[LibrarySequence, ...]
    parse_failures: tuple[ParseFailure, ...]
    exact_pairs: tuple[ExactPair, ...]
    best_matches: tuple[BestMatch, ...]
    internal_duplicate_groups: tuple[DuplicateGroup, ...]
    public_duplicate_groups: tuple[DuplicateGroup, ...]


def _category(relative_path: Path) -> str:
    return relative_path.parts[0] if len(relative_path.parts) > 1 else "根目录"


def _sequence_hash(sequence: str) -> str:
    canonical = canonical_circular_sequence(sequence)
    return hashlib.sha256(canonical.encode("ascii")).hexdigest()


def _scan_library(
    root: Path,
    *,
    library: str,
    extensions: set[str],
    k: int,
    sketch_size: int,
) -> tuple[list[LibrarySequence], list[ParseFailure]]:
    records: list[LibrarySequence] = []
    failures: list[ParseFailure] = []
    normalized_extensions = {suffix.lower() for suffix in extensions}
    paths = sorted(
        (
            path
            for path in root.rglob("*")
            if path.is_file() and path.suffix.lower() in normalized_extensions
        ),
        key=lambda item: str(item).casefold(),
    )
    for path in paths:
        try:
            parsed_records = read_sequence_file(path)
            relative = path.relative_to(root)
            for parsed in parsed_records:
                file_name = path.name
                if len(parsed_records) > 1:
                    file_name = f"{path.name}#{parsed.record_index}"
                sequence = parsed.sequence
                records.append(
                    LibrarySequence(
                        library=library,
                        file_name=file_name,
                        display_name=parsed.display_name,
                        path=path,
                        relative_path=str(relative),
                        category=_category(relative),
                        role=(
                            infer_internal_sequence_role(relative)
                            if library == "内部库"
                            else "公共载体图谱"
                        ),
                        sequence=sequence,
                        format_name=parsed.format_name,
                        topology=parsed.topology,
                        feature_labels=parsed.feature_labels,
                        ambiguous_base_count=sum(base not in "ACGT" for base in sequence),
                        canonical_hash=_sequence_hash(sequence),
                        sketch=canonical_kmer_sketch(sequence, k=k, sketch_size=sketch_size),
                    )
                )
        except Exception as exc:  # keep one malformed file from blocking the inventory
            failures.append(ParseFailure(library=library, path=path, error=str(exc)))
    return records, failures


def _group_by_hash(records: Iterable[LibrarySequence]):
    groups: dict[str, list[LibrarySequence]] = defaultdict(list)
    for record in records:
        groups[record.canonical_hash].append(record)
    return groups


def _duplicate_groups(
    groups: dict[str, list[LibrarySequence]],
    library: str,
) -> list[DuplicateGroup]:
    return [
        DuplicateGroup(library, sequence_hash, tuple(records))
        for sequence_hash, records in groups.items()
        if len(records) > 1
    ]


def _normalized_name(record: LibrarySequence) -> str:
    text = f"{record.path.stem} {record.display_name}".casefold()
    return "".join(character for character in text if character.isalnum())


def _name_similarity(left: LibrarySequence, right: LibrarySequence) -> float:
    return SequenceMatcher(None, _normalized_name(left), _normalized_name(right)).ratio()


def _sketch_similarity(left: tuple[int, ...], right: tuple[int, ...]) -> float:
    left_set = set(left)
    right_set = set(right)
    union = left_set | right_set
    if not union:
        return 0.0
    return len(left_set & right_set) / len(union) * 100


def _candidate_indices(
    internal: LibrarySequence,
    public_records: list[LibrarySequence],
    inverted_index: dict[int, list[int]],
    candidate_count: int,
) -> list[int]:
    overlaps: Counter[int] = Counter()
    for kmer in internal.sketch:
        overlaps.update(inverted_index.get(kmer, ()))

    ranked = sorted(
        overlaps,
        key=lambda index: (
            overlaps[index],
            _name_similarity(internal, public_records[index]),
            -abs(internal.length - public_records[index].length),
        ),
        reverse=True,
    )
    selected = ranked[:candidate_count]

    name_candidates = sorted(
        range(len(public_records)),
        key=lambda index: _name_similarity(internal, public_records[index]),
        reverse=True,
    )[:2]
    length_candidates = sorted(
        range(len(public_records)),
        key=lambda index: abs(internal.length - public_records[index].length),
    )[:2]
    for index in (*name_candidates, *length_candidates):
        if index not in selected:
            selected.append(index)
    return selected


def analyze_vector_libraries(
    internal_root: Path,
    public_root: Path,
    *,
    internal_extensions: set[str] | None = None,
    public_extensions: set[str] | None = None,
    k: int = 17,
    sketch_size: int = 256,
    candidate_count: int = 4,
) -> VectorComparisonReport:
    internal_root = Path(internal_root).resolve()
    public_root = Path(public_root).resolve()
    internal_records, internal_failures = _scan_library(
        internal_root,
        library="内部库",
        extensions=internal_extensions or DEFAULT_INTERNAL_EXTENSIONS,
        k=k,
        sketch_size=sketch_size,
    )
    public_records, public_failures = _scan_library(
        public_root,
        library="公共库",
        extensions=public_extensions or DEFAULT_PUBLIC_EXTENSIONS,
        k=k,
        sketch_size=sketch_size,
    )
    if not internal_records:
        raise ValueError(f"No internal sequences could be read from {internal_root}")
    if not public_records:
        raise ValueError(f"No public sequences could be read from {public_root}")

    internal_groups = _group_by_hash(internal_records)
    public_groups = _group_by_hash(public_records)
    exact_pairs = [
        ExactPair(internal, public)
        for sequence_hash, internals in internal_groups.items()
        for internal in internals
        for public in public_groups.get(sequence_hash, ())
    ]

    inverted_index: dict[int, list[int]] = defaultdict(list)
    for index, public in enumerate(public_records):
        for kmer in public.sketch:
            inverted_index[kmer].append(index)

    best_matches: list[BestMatch] = []
    for internal in internal_records:
        exact_aliases = tuple(public_groups.get(internal.canonical_hash, ()))
        if exact_aliases:
            public = max(exact_aliases, key=lambda item: _name_similarity(internal, item))
            comparison = compare_circular_sequences(internal.sequence, public.sequence)
            similarity = 100.0
        else:
            candidate_indices = _candidate_indices(
                internal,
                public_records,
                inverted_index,
                candidate_count,
            )
            evaluated = []
            for index in candidate_indices:
                public_candidate = public_records[index]
                comparison_candidate = compare_circular_sequences(
                    internal.sequence,
                    public_candidate.sequence,
                )
                evaluated.append(
                    (
                        comparison_candidate.identity_percent,
                        -comparison_candidate.edit_distance,
                        _sketch_similarity(internal.sketch, public_candidate.sketch),
                        _name_similarity(internal, public_candidate),
                        public_candidate,
                        comparison_candidate,
                    )
                )
            _, _, similarity, _, public, comparison = max(evaluated, key=lambda item: item[:4])

        best_matches.append(
            BestMatch(
                internal=internal,
                public=public,
                comparison=comparison,
                classification=classify_similarity(
                    comparison.identity_percent,
                    comparison.edit_distance,
                ),
                sketch_similarity_percent=round(similarity, 4),
                exact_public_aliases=exact_aliases,
            )
        )

    return VectorComparisonReport(
        internal_root=internal_root,
        public_root=public_root,
        generated_at=datetime.now(),
        internal_records=tuple(internal_records),
        public_records=tuple(public_records),
        parse_failures=tuple(internal_failures + public_failures),
        exact_pairs=tuple(exact_pairs),
        best_matches=tuple(best_matches),
        internal_duplicate_groups=tuple(_duplicate_groups(internal_groups, "内部库")),
        public_duplicate_groups=tuple(_duplicate_groups(public_groups, "公共库")),
    )


def _annotation_difference(left: LibrarySequence, right: LibrarySequence) -> str:
    left_labels = set(left.feature_labels)
    right_labels = set(right.feature_labels)
    internal_only = sorted(left_labels - right_labels)
    public_only = sorted(right_labels - left_labels)
    parts = []
    if internal_only:
        parts.append("内部特有注释：" + "; ".join(internal_only[:10]))
    if public_only:
        parts.append("公共特有注释：" + "; ".join(public_only[:10]))
    return "；".join(parts) or "注释名称集合一致"


def _difference_kind(kind: str) -> str:
    return {
        "replace": "替换",
        "delete": "内部多出",
        "insert": "内部缺失",
    }.get(kind, kind)


def _position_text(start: int, end: int) -> str:
    if start == end:
        return f"{start + 1} 位点间"
    return f"{start + 1}-{end}"


def _truncate_sequence(sequence: str, limit: int = 200) -> str:
    if len(sequence) <= limit:
        return sequence
    return sequence[:limit] + f"...（共 {len(sequence)} bp）"


def _best_match_rows(report: VectorComparisonReport) -> list[list[object]]:
    rows = []
    for match in report.best_matches:
        comparison = match.comparison
        rows.append(
            [
                match.internal.file_name,
                match.internal.display_name,
                match.internal.role,
                match.internal.category,
                str(match.internal.path),
                match.internal.length,
                match.internal.topology,
                match.public.file_name,
                match.public.display_name,
                match.public.category,
                str(match.public.path),
                match.public.length,
                match.classification,
                comparison.identity_percent,
                comparison.edit_distance,
                comparison.substitutions,
                comparison.internal_extra_bp,
                comparison.internal_missing_bp,
                comparison.difference_block_count,
                match.internal.length - match.public.length,
                comparison.public_orientation,
                comparison.public_rotation,
                match.sketch_similarity_percent,
                len(match.exact_public_aliases),
                "\n".join(alias.file_name for alias in match.exact_public_aliases),
                "; ".join(match.internal.feature_labels),
                "; ".join(match.public.feature_labels),
                _annotation_difference(match.internal, match.public),
                "完全一致仍需核对载体注释和实际批次"
                if comparison.edit_distance == 0
                else "请人工复核差异区段及载体用途",
            ]
        )
    return rows


BEST_MATCH_HEADERS = [
    "内部文件名",
    "内部载体名",
    "内部序列类型",
    "内部分类目录",
    "内部存储位置",
    "内部长度_bp",
    "内部拓扑",
    "公共库最佳文件名",
    "公共载体名",
    "公共库分类",
    "公共库存储位置",
    "公共长度_bp",
    "序列结论",
    "序列一致度_%",
    "编辑距离_bp",
    "替换_bp",
    "内部多出_bp",
    "内部缺失_bp",
    "差异区段数",
    "长度差_内部减公共_bp",
    "公共序列方向",
    "公共序列旋转起点_0基",
    "kmer草图相似度_%",
    "完全一致公共别名数",
    "完全一致公共别名",
    "内部feature摘要",
    "公共feature摘要",
    "注释差异摘要",
    "复核提示",
]


def _exact_rows(report: VectorComparisonReport) -> list[list[object]]:
    rows = []
    for pair in report.exact_pairs:
        comparison = compare_circular_sequences(pair.internal.sequence, pair.public.sequence)
        rows.append(
            [
                pair.internal.file_name,
                pair.internal.display_name,
                pair.internal.role,
                str(pair.internal.path),
                pair.internal.length,
                pair.public.file_name,
                pair.public.display_name,
                pair.public.category,
                str(pair.public.path),
                pair.public.length,
                comparison.public_orientation,
                comparison.public_rotation,
                "; ".join(pair.internal.feature_labels),
                "; ".join(pair.public.feature_labels),
                _annotation_difference(pair.internal, pair.public),
            ]
        )
    return rows


def _difference_rows(report: VectorComparisonReport) -> list[list[object]]:
    rows = []
    for match in report.best_matches:
        if match.classification == "未发现可靠近似匹配":
            continue
        for block_number, block in enumerate(match.comparison.difference_blocks, start=1):
            rows.append(
                [
                    match.internal.file_name,
                    str(match.internal.path),
                    match.public.file_name,
                    str(match.public.path),
                    match.classification,
                    match.comparison.identity_percent,
                    block_number,
                    _difference_kind(block.kind),
                    _position_text(block.internal_start, block.internal_end),
                    _position_text(block.public_start, block.public_end),
                    block.internal_end - block.internal_start,
                    block.public_end - block.public_start,
                    _truncate_sequence(block.internal_sequence),
                    _truncate_sequence(block.public_sequence),
                    "是" if match.comparison.difference_blocks_truncated else "否",
                ]
            )
    return rows


def _duplicate_rows(groups: tuple[DuplicateGroup, ...]) -> list[list[object]]:
    rows = []
    for group_number, group in enumerate(groups, start=1):
        for record in group.records:
            rows.append(
                [
                    group_number,
                    len(group.records),
                    record.file_name,
                    record.display_name,
                    record.role,
                    record.length,
                    str(record.path),
                    group.canonical_hash,
                ]
            )
    return rows


def _inventory_rows(records: tuple[LibrarySequence, ...]) -> list[list[object]]:
    return [
        [
            record.file_name,
            record.display_name,
            record.role,
            record.category,
            record.length,
            record.topology,
            record.ambiguous_base_count,
            record.format_name,
            str(record.path),
            "; ".join(record.feature_labels),
            record.canonical_hash,
        ]
        for record in records
    ]


def _add_table_sheet(
    workbook: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[object]],
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for column_index, header in enumerate(headers, start=1):
        width = 16
        if "位置" in header or "摘要" in header or "feature" in header:
            width = 45
        elif "名称" in header or "文件名" in header or "别名" in header:
            width = 28
        elif "序列" in header:
            width = 36
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if title == "内部载体最佳匹配":
        conclusion_column = headers.index("序列结论") + 1
        fills = {
            "完全一致": "C6EFCE",
            "高度一致（小改造）": "E2F0D9",
            "同骨架高度相似": "FFF2CC",
            "部分相似": "FCE4D6",
            "未发现可靠近似匹配": "F4CCCC",
        }
        for row_number in range(2, sheet.max_row + 1):
            value = sheet.cell(row_number, conclusion_column).value
            color = fills.get(value)
            if color:
                for cell in sheet[row_number]:
                    cell.fill = PatternFill("solid", fgColor=color)


def _summary_rows(report: VectorComparisonReport) -> list[list[object]]:
    classifications = Counter(match.classification for match in report.best_matches)
    return [
        ["生成时间", report.generated_at.strftime("%Y-%m-%d %H:%M:%S")],
        ["内部库根目录", str(report.internal_root)],
        ["公共库根目录", str(report.public_root)],
        ["内部可解析序列数", len(report.internal_records)],
        ["其中内部载体图谱数", sum(r.role == "载体图谱" for r in report.internal_records)],
        ["其中测序共识/FASTA参考数", sum(r.role != "载体图谱" for r in report.internal_records)],
        ["公共可解析序列数", len(report.public_records)],
        ["跨库完全一致配对数", len(report.exact_pairs)],
        ["有至少一个完全一致公共载体的内部序列数", sum(bool(m.exact_public_aliases) for m in report.best_matches)],
        ["内部重复序列组数", len(report.internal_duplicate_groups)],
        ["公共重复序列组数", len(report.public_duplicate_groups)],
        ["解析失败文件数", len(report.parse_failures)],
        ["完全一致", classifications["完全一致"]],
        ["高度一致（小改造）", classifications["高度一致（小改造）"]],
        ["同骨架高度相似", classifications["同骨架高度相似"]],
        ["部分相似", classifications["部分相似"]],
        ["未发现可靠近似匹配", classifications["未发现可靠近似匹配"]],
        ["完全一致定义", "环状序列允许起始坐标旋转和整条反向互补后，编辑距离仍为 0"],
        ["高度一致定义", "序列一致度 >= 99.5%，但存在至少 1 bp 差异"],
        ["同骨架高度相似定义", "序列一致度 >= 95% 且 < 99.5%"],
        ["部分相似定义", "序列一致度 >= 80% 且 < 95%"],
        ["差异区段展开规则", "只展开序列一致度 >= 80% 的最佳候选；低相似候选仅保留汇总数值"],
        ["重要限制", "结果是序列层面的候选关系；载体用途、标签、抗性和实物批次仍需人工核对"],
    ]


def _write_main_csv(output: Path, rows: list[list[object]]) -> None:
    with output.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(BEST_MATCH_HEADERS)
        writer.writerows(rows)


def write_vector_comparison_report(
    report: VectorComparisonReport,
    output_path: Path,
    *,
    main_csv_path: Path | None = None,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    _add_table_sheet(workbook, "总览", ["项目", "结果"], _summary_rows(report))
    _add_table_sheet(
        workbook,
        "完全一致_跨库",
        [
            "内部文件名",
            "内部载体名",
            "内部序列类型",
            "内部存储位置",
            "内部长度_bp",
            "公共文件名",
            "公共载体名",
            "公共分类",
            "公共存储位置",
            "公共长度_bp",
            "公共序列方向",
            "公共序列旋转起点_0基",
            "内部feature摘要",
            "公共feature摘要",
            "注释差异摘要",
        ],
        _exact_rows(report),
    )
    best_rows = _best_match_rows(report)
    _add_table_sheet(workbook, "内部载体最佳匹配", BEST_MATCH_HEADERS, best_rows)
    _add_table_sheet(
        workbook,
        "差异区段",
        [
            "内部文件名",
            "内部存储位置",
            "公共文件名",
            "公共存储位置",
            "序列结论",
            "序列一致度_%",
            "区段序号",
            "差异类型",
            "内部位置_1基",
            "公共位置_1基",
            "内部区段长度_bp",
            "公共区段长度_bp",
            "内部序列",
            "公共序列",
            "区段是否截断未完全列出",
        ],
        _difference_rows(report),
    )
    duplicate_headers = [
        "重复组号",
        "组内文件数",
        "文件名",
        "载体名",
        "序列类型",
        "长度_bp",
        "存储位置",
        "环状标准化SHA256",
    ]
    _add_table_sheet(
        workbook,
        "内部重复序列",
        duplicate_headers,
        _duplicate_rows(report.internal_duplicate_groups),
    )
    _add_table_sheet(
        workbook,
        "公共重复序列",
        duplicate_headers,
        _duplicate_rows(report.public_duplicate_groups),
    )
    inventory_headers = [
        "文件名",
        "载体名",
        "序列类型",
        "分类目录",
        "长度_bp",
        "拓扑",
        "模糊碱基数",
        "文件格式",
        "存储位置",
        "feature摘要",
        "环状标准化SHA256",
    ]
    _add_table_sheet(
        workbook,
        "内部载体清单",
        inventory_headers,
        _inventory_rows(report.internal_records),
    )
    _add_table_sheet(
        workbook,
        "公共载体清单",
        inventory_headers,
        _inventory_rows(report.public_records),
    )
    _add_table_sheet(
        workbook,
        "解析失败",
        ["库", "文件位置", "错误"],
        [[failure.library, str(failure.path), failure.error] for failure in report.parse_failures],
    )
    workbook.save(output_path)
    if main_csv_path is not None:
        main_csv_path = Path(main_csv_path)
        main_csv_path.parent.mkdir(parents=True, exist_ok=True)
        _write_main_csv(main_csv_path, best_rows)
    return output_path
